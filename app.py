"""
INFY Migration Version Lifecycle Tracker
5-Agent Streamlit Application — Infosys Enterprise Architecture

Core design:
- baseline_data.py holds ALL 321 OS/DB rows from AI training knowledge
- Data is visible IMMEDIATELY on app load — no waiting for Agent 1
- Agent 1 checks internet for date changes and updates Notes column
- Agent 2 fills Recommendation column via OpenAI
- Agent 3 refreshes every 14 days with permission
- Agent 4 snapshots data before every refresh
- Agent 5 interactive policy analysis → guiding principles → verdicts

Project window: Apr 2026 → Jun 2028
"""

import streamlit as st
import pandas as pd
import json
from datetime import datetime
import sys, os, types, importlib.util

# ── Python 3.14 Streamlit Cloud safe loader ───────────────────────────────────
_HERE = os.path.abspath(os.path.dirname(__file__))

def _load(dotted_name: str, rel_path: str):
    cached = sys.modules.get(dotted_name)
    if cached is not None and getattr(cached, "__file__", None):
        return cached
    abs_path = os.path.join(_HERE, rel_path)
    spec     = importlib.util.spec_from_file_location(dotted_name, abs_path)
    mod      = importlib.util.module_from_spec(spec)
    mod.__file__ = abs_path
    sys.modules[dotted_name] = mod
    try:
        spec.loader.exec_module(mod)
    except Exception:
        sys.modules.pop(dotted_name, None)
        raise
    return mod

for _pkg in ("agents", "utils"):
    if _pkg not in sys.modules:
        _ns = types.ModuleType(_pkg)
        _ns.__path__ = [os.path.join(_HERE, _pkg)]
        _ns.__package__ = _pkg
        sys.modules[_pkg] = _ns

_m_baseline  = _load("baseline_data",         "baseline_data.py")
_m_os        = _load("agents.agent_os",        "agents/agent_os.py")
_m_db        = _load("agents.agent_db",        "agents/agent_db.py")
_m_refresh   = _load("agents.agent_refresh",   "agents/agent_refresh.py")
_m_version   = _load("agents.agent_versioning","agents/agent_versioning.py")
_m_analysis  = _load("agents.agent_analysis",  "agents/agent_analysis.py")
_m_export    = _load("utils.excel_export",     "utils/excel_export.py")
_m_store     = _load("utils.data_store",       "utils/data_store.py")
_m_risk      = _load("utils.risk_scoring",     "utils/risk_scoring.py")
_m_dashboard = _load("utils.dashboard",        "utils/dashboard.py")
_m_config    = _load("utils.config",           "utils/config.py")
_m_inventory = _load("utils.inventory_upload",  "utils/inventory_upload.py")
_m_scenario  = _load("utils.scenario_planner",  "utils/scenario_planner.py")
_m_pdf       = _load("utils.pdf_report",        "utils/pdf_report.py")

OSDataAgent          = _m_os.OSDataAgent
RecommendationAgent  = _m_db.RecommendationAgent
RefreshAgent         = _m_refresh.RefreshAgent
VersionGuardianAgent = _m_version.VersionGuardianAgent
PolicyAnalysisAgent  = _m_analysis.PolicyAnalysisAgent
export_to_excel      = _m_export.export_to_excel
OS_DATA              = _m_baseline.OS_DATA
DB_DATA              = _m_baseline.DB_DATA
OS_COLUMNS           = _m_baseline.OS_COLUMNS
DB_COLUMNS           = _m_baseline.DB_COLUMNS
WS_DATA              = _m_baseline.WS_DATA
WS_COLUMNS           = _m_baseline.WS_COLUMNS
AS_DATA              = _m_baseline.AS_DATA
AS_COLUMNS           = _m_baseline.AS_COLUMNS
FW_DATA              = _m_baseline.FW_DATA
FW_COLUMNS           = _m_baseline.FW_COLUMNS

# Data store helpers
save_os_df      = _m_store.save_os_df
save_db_df      = _m_store.save_db_df
load_os_df      = _m_store.load_os_df
load_db_df      = _m_store.load_db_df
save_app_state  = _m_store.save_app_state
load_app_state  = _m_store.load_app_state
has_stored_data = _m_store.has_stored_data
get_rec_stats   = _m_store.get_rec_stats

# New enhancement modules
add_risk_scores       = _m_risk.add_risk_scores
get_risk_summary      = _m_risk.get_risk_summary
risk_distribution_chart = _m_dashboard.risk_distribution_chart
status_breakdown_chart  = _m_dashboard.status_breakdown_chart
eol_timeline_chart      = _m_dashboard.eol_timeline_chart
top_urgent_items        = _m_dashboard.top_urgent_items
risk_score_histogram    = _m_dashboard.risk_score_histogram
render_project_settings = _m_config.render_project_settings
get_project_end         = _m_config.get_project_end
render_upload_section   = _m_inventory.render_upload_section
match_os_inventory      = _m_inventory.match_os_inventory
match_db_inventory      = _m_inventory.match_db_inventory
render_inventory_results = _m_inventory.render_inventory_results
render_scenario_planner  = _m_scenario.render_scenario_planner
generate_pdf_report      = _m_pdf.generate_pdf_report

# ── Page config ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="INFY Version Lifecycle Tracker",
    page_icon="🖥️",
    layout="wide",
    initial_sidebar_state="expanded",
    menu_items={"About": "INFY Migration Version Lifecycle Tracker — Powered by OpenAI GPT"}
)

# ── CSS ───────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
.infy-header {
    background: linear-gradient(135deg, #001F5B 0%, #003087 60%, #0057C8 100%);
    padding: 1.2rem 2rem; border-radius: 12px;
    color: white; margin-bottom: 1.5rem;
}
.infy-header h1 { margin: 0; font-size: 1.55rem; font-weight: 700; }
.infy-header p  { margin: 4px 0 0; font-size: 0.82rem; opacity: 0.8; }
.agent-card {
    border-left: 4px solid; border-radius: 0 8px 8px 0;
    padding: 0.7rem 1rem; margin-bottom: 0.65rem;
    background: rgba(0,0,0,0.025);
}
.a1{border-color:#0EA5E9;} .a2{border-color:#8B5CF6;}
.a3{border-color:#F59E0B;} .a4{border-color:#10B981;} .a5{border-color:#7C3AED;}
.badge {
    display: inline-block; padding: 2px 8px; border-radius: 12px;
    font-size: 0.68rem; font-weight: 600; text-transform: uppercase;
}
.b-idle    { background:#F1F5F9; color:#64748B; }
.b-running { background:#FEF3C7; color:#92400E; }
.b-done    { background:#D1FAE5; color:#065F46; }
table td, table th { word-wrap: break-word; overflow-wrap: break-word; }
.stTabs [data-baseweb="tab-list"] {
    justify-content: center;
}
.b-error   { background:#FEE2E2; color:#991B1B; }

/* ── Sidebar navigation ─────────────────────────────────────────── */
.nav-section-header {
    font-size: 0.7rem; font-weight: 700; color: #94A3B8;
    letter-spacing: 1.5px; padding: 12px 12px 4px; margin-top: 6px;
}
.nav-item {
    display: flex; align-items: center; gap: 10px;
    padding: 8px 14px; border-radius: 8px; cursor: pointer;
    font-size: 0.85rem; color: #334155; transition: all 0.2s;
    margin: 1px 4px; text-decoration: none;
}
.nav-item:hover { background: #EFF6FF; color: #1D4ED8; }
.nav-item.active {
    background: linear-gradient(135deg, #DBEAFE, #EFF6FF);
    color: #1E40AF; font-weight: 600;
    border-left: 3px solid #3B82F6;
}
.nav-item .nav-icon { font-size: 1rem; width: 22px; text-align: center; }
.nav-item .nav-dot {
    width: 8px; height: 8px; border-radius: 50%;
    background: #60A5FA; margin-left: auto; flex-shrink: 0;
}

/* ── Flowchart ───────────────────────────────────────────────────── */
.flowchart-container {
    display: flex; flex-wrap: wrap; align-items: center;
    justify-content: center; gap: 8px; padding: 1.2rem;
    background: linear-gradient(135deg, #0F172A 0%, #1E293B 100%);
    border-radius: 12px; margin-bottom: 1.5rem;
}
.fc-block {
    padding: 10px 16px; border-radius: 8px; font-size: 0.78rem;
    font-weight: 600; cursor: pointer; transition: all 0.3s;
    text-align: center; min-width: 100px; position: relative;
    border: 2px solid transparent;
}
.fc-block.fc-idle {
    background: #334155; color: #94A3B8; border-color: #475569;
}
.fc-block.fc-active {
    background: #1D4ED8; color: white; border-color: #60A5FA;
    box-shadow: 0 0 12px rgba(59,130,246,0.5);
    animation: fc-pulse 2s ease-in-out infinite;
}
.fc-block.fc-done {
    background: #065F46; color: #6EE7B7; border-color: #10B981;
}
.fc-block:hover {
    transform: translateY(-2px);
    box-shadow: 0 4px 12px rgba(0,0,0,0.3);
}
.fc-arrow {
    color: #64748B; font-size: 1.2rem; font-weight: 700;
}
@keyframes fc-pulse {
    0%, 100% { box-shadow: 0 0 12px rgba(59,130,246,0.5); }
    50% { box-shadow: 0 0 24px rgba(59,130,246,0.8); }
}

/* ── Blinking AI Chat button ─────────────────────────────────────── */
.ai-chat-btn {
    display: inline-flex; align-items: center; gap: 8px;
    padding: 12px 24px; border-radius: 25px; font-size: 0.95rem;
    font-weight: 700; cursor: pointer; border: none;
    background: linear-gradient(135deg, #7C3AED, #9333EA);
    color: white; transition: all 0.3s;
    animation: chat-blink 1.5s ease-in-out infinite;
}
.ai-chat-btn:hover {
    transform: scale(1.05);
    box-shadow: 0 6px 20px rgba(124,58,237,0.5);
}
@keyframes chat-blink {
    0%, 100% { box-shadow: 0 0 8px rgba(124,58,237,0.4); }
    50% { box-shadow: 0 0 20px rgba(124,58,237,0.8), 0 0 40px rgba(124,58,237,0.3); }
}
.chat-dot {
    width: 10px; height: 10px; border-radius: 50%;
    background: #34D399; animation: dot-blink 1s infinite;
}
@keyframes dot-blink {
    0%, 100% { opacity: 1; }
    50% { opacity: 0.3; }
}

/* ── GP Category Cards ───────────────────────────────────────────── */
.gp-card {
    background: white; border-radius: 12px; padding: 1.2rem;
    border: 1px solid #E2E8F0; transition: all 0.3s;
    text-align: center; cursor: default;
}
.gp-card:hover {
    transform: translateY(-4px);
    box-shadow: 0 8px 24px rgba(0,0,0,0.1);
}
.gp-card .gp-icon { font-size: 2rem; margin-bottom: 8px; }
.gp-card .gp-title { font-size: 0.95rem; font-weight: 700; color: #1E293B; }
.gp-card .gp-desc { font-size: 0.78rem; color: #64748B; margin-top: 4px; }
</style>
""", unsafe_allow_html=True)


# ── Session state ─────────────────────────────────────────────────────────────
def _init():
    if "os_df" not in st.session_state:
        # ── Load from SQLite if data exists, else load baseline and save it ──
        if has_stored_data():
            stored_os = load_os_df()
            stored_db = load_db_df()
            st.session_state["os_df"] = stored_os
            st.session_state["db_df"] = stored_db
        else:
            # First ever run — load baseline and immediately persist to SQLite
            baseline_os = pd.DataFrame(OS_DATA)
            baseline_db = pd.DataFrame(DB_DATA)
            save_os_df(baseline_os)
            save_db_df(baseline_db)
            st.session_state["os_df"] = baseline_os
            st.session_state["db_df"] = baseline_db

    # ── Web Servers, App Servers, Frameworks — always from baseline ──
    if "ws_df" not in st.session_state:
        st.session_state["ws_df"] = pd.DataFrame(WS_DATA)
    if "as_df" not in st.session_state:
        st.session_state["as_df"] = pd.DataFrame(AS_DATA)
    if "fw_df" not in st.session_state:
        st.session_state["fw_df"] = pd.DataFrame(FW_DATA)

    # Restore last_refresh from SQLite so Agent 3 timer survives restarts
    if "last_refresh" not in st.session_state:
        saved_refresh = load_app_state("last_refresh")
        if saved_refresh:
            try:
                st.session_state["last_refresh"] = datetime.fromisoformat(saved_refresh)
            except Exception:
                st.session_state["last_refresh"] = None
        else:
            st.session_state["last_refresh"] = None

    # Other session defaults
    for k, v in {
        "a1_status": "idle", "a2_status": "idle",
        "changes_log": [], "old_os_df": None, "old_db_df": None,
        "a3_skip_until": None,
        "current_page": "Discovery",
    }.items():
        if k not in st.session_state:
            st.session_state[k] = v

_init()
refresh_agent = RefreshAgent()
VersionGuardianAgent.init_session()
PolicyAnalysisAgent.init_session()

# ── Restore Strategist survey data from SQLite if session is fresh ────────
_strategist_keys = [
    "a5_status", "a5_landscape_selected", "a5_db_landscape_selected",
    "a5_context", "a5_principles_table_data", "a5_costed_data",
    "a5_custom_cloud_profiles",
]
if st.session_state.get("a5_status") == "idle":
    saved_a5 = load_app_state("strategist_survey")
    if saved_a5 and isinstance(saved_a5, dict):
        for k in _strategist_keys:
            if k in saved_a5 and saved_a5[k]:
                st.session_state[k] = saved_a5[k]
        # Safety: if restored status is beyond survey but no landscape was selected,
        # reset to survey so user must fill in the forms first
        _restored_status = st.session_state.get("a5_status", "idle")
        _has_landscape = bool(st.session_state.get("a5_landscape_selected"))
        if _restored_status not in ("idle", "survey") and not _has_landscape:
            st.session_state["a5_status"] = "idle"


def _save_strategist_state():
    """Persist all Strategist survey data to SQLite."""
    data = {}
    for k in _strategist_keys:
        val = st.session_state.get(k)
        if val is not None:
            data[k] = val
    if data:
        save_app_state("strategist_survey", data)

# ── Compute risk scores on every load ─────────────────────────────────────
if "Risk Score" not in st.session_state.os_df.columns:
    st.session_state.os_df = add_risk_scores(st.session_state.os_df, "OS")
if "Risk Score" not in st.session_state.db_df.columns:
    st.session_state.db_df = add_risk_scores(st.session_state.db_df, "DB")
if "Risk Score" not in st.session_state.ws_df.columns:
    st.session_state.ws_df = add_risk_scores(st.session_state.ws_df, "DB")
if "Risk Score" not in st.session_state.as_df.columns:
    st.session_state.as_df = add_risk_scores(st.session_state.as_df, "DB")
if "Risk Score" not in st.session_state.fw_df.columns:
    st.session_state.fw_df = add_risk_scores(st.session_state.fw_df, "DB")


def badge(status: str) -> str:
    icons = {"idle":"◌","running":"⟳","done":"✓","error":"✕"}
    cls   = {"idle":"idle","running":"running","done":"done","error":"error"}.get(status,"idle")
    return f'<span class="badge b-{cls}">{icons.get(cls,"·")} {status.upper()}</span>'


# ── Sidebar ────────────────────────────────────────────────────────────────────

# Navigation page definitions
_NAV_PAGES = {
    "DAY 0 — ASSESS": [
        ("Discovery",              "🔍", True),
        ("Disposition",            "📋", True),
        ("Future Blueprint",       "📐", True),
        ("Workload Optimization",  "🔧", True),
        ("Business Case",          "💼", True),
        ("Server Affinity",        "🖧", True),
        ("Code Analysis",          "💻", True),
        ("Multi-Cloud Strategy",   "☁️", True),
        ("Version Lifecycle",      "⏱️", True),
        ("Wave Planning",          "🌊", True),
        ("AI Copilot — Assess",    "🤖", True),
    ],
    "DAY 1 — MIGRATE": [
        ("AI Migration Hub",       "🚀", True),
    ],
    "DAY 2 — OPERATE": [
        ("Compliance",             "🛡️", True),
        ("Governance",             "⚖️", True),
        ("Operate & Monitor",      "📡", True),
        ("FinOps",                 "💲", True),
        ("License Optimization",   "📄", True),
        ("PaaS / Serverless",      "🔗", True),
        ("Dependency Discovery",   "🔎", True),
        ("AI Copilot — Operate",   "🤖", True),
    ],
    "OPTIMIZE": [
        ("Cloud Modernization",    "☁️", True),
    ],
}

with st.sidebar:
    st.markdown("""
    <div style="background:linear-gradient(135deg,#001F5B 0%,#003087 50%,#0057C8 100%);
                border-radius:12px;padding:14px 16px;text-align:center;">
      <div style="display:flex;align-items:center;justify-content:center;gap:10px;margin-bottom:6px;">
        <div style="position:relative;width:40px;height:40px;">
          <svg viewBox="0 0 40 40" width="40" height="40" style="animation:logospin 8s linear infinite;">
            <circle cx="20" cy="20" r="16" fill="none" stroke="#00C6FF" stroke-width="2" opacity="0.3"/>
            <circle cx="20" cy="20" r="12" fill="none" stroke="#00C6FF" stroke-width="2"
                    stroke-dasharray="18 58" stroke-linecap="round"/>
          </svg>
          <div style="position:absolute;top:50%;left:50%;transform:translate(-50%,-50%);
                      width:6px;height:6px;background:#00C6FF;border-radius:50%;"></div>
        </div>
        <div style="text-align:left;">
          <div style="font-size:20px;font-weight:800;color:white;letter-spacing:2px;line-height:1;">INFY</div>
          <div style="font-size:8.5px;color:#93C5FD;letter-spacing:2.5px;margin-top:3px;">LIFECYCLE TRACKER</div>
        </div>
      </div>
      <div style="height:2px;background:#1E3A5F;border-radius:1px;overflow:hidden;margin:4px 0;">
        <div style="height:100%;background:linear-gradient(90deg,#10B981,#34D399,#6EE7B7);
                    border-radius:1px;animation:logosweep 4s ease-in-out infinite;"></div>
      </div>
      <div style="display:flex;justify-content:center;gap:12px;margin-top:6px;">
        <span style="font-size:7.5px;color:#6EE7B7;display:flex;align-items:center;gap:3px;">
          <span style="width:5px;height:5px;background:#10B981;border-radius:50%;display:inline-block;"></span>LIVE</span>
        <span style="font-size:7.5px;color:#FCD34D;display:flex;align-items:center;gap:3px;">
          <span style="width:5px;height:5px;background:#F59E0B;border-radius:50%;display:inline-block;"></span>5 AGENTS</span>
        <span style="font-size:7.5px;color:#93C5FD;display:flex;align-items:center;gap:3px;">
          <span style="width:5px;height:5px;background:#60A5FA;border-radius:50%;display:inline-block;"></span>AI</span>
      </div>
    </div>
    <style>
      @keyframes logospin{from{transform:rotate(0deg)}to{transform:rotate(360deg)}}
      @keyframes logosweep{0%,100%{width:20%;margin-left:0}50%{width:80%;margin-left:10%}}
    </style>
    """, unsafe_allow_html=True)

    # ── AI Advisor button (always visible at top of sidebar) ─────────────────
    _sb_chat_open = st.session_state.get("show_chat", False)
    _sb_a5s = st.session_state.get("a5_status", "idle")
    if _sb_chat_open:
        if st.button("✕ Close AI Chat", key="close_chat_sb", use_container_width=True):
            st.session_state["show_chat"] = False
            st.rerun()
    else:
        if st.button("🧠 Open AI Advisor", key="open_chat_sb",
                     use_container_width=True, type="primary"):
            st.session_state["show_chat"] = True
            if st.session_state.get("a5_status") in ("idle",):
                st.session_state["a5_status"] = "survey"
            st.rerun()
    if _sb_a5s not in ("idle",):
        if st.button("🔄 Start Fresh", key="reset_btn_sb", use_container_width=True):
            for _k in ["a5_status", "a5_landscape_selected", "a5_db_landscape_selected",
                        "a5_context", "a5_principles_table_data", "a5_costed_data",
                        "a5_custom_cloud_profiles", "a5_principles", "a5_costs",
                        "a5_preflight_done", "a5_os_done", "a5_db_done",
                        "a5_ws_done", "a5_as_done", "a5_fw_done",
                        "a5_session_id", "a5_opening_pending", "a5_log",
                        "show_chat", "show_strategist"]:
                st.session_state.pop(_k, None)
            st.session_state["a5_status"] = "idle"
            st.session_state["show_chat"] = False
            try:
                save_app_state("strategist_survey", {})
            except Exception:
                pass
            st.rerun()
    st.divider()

    # ── API Key (compact) ─────────────────────────────────────────────────────
    with st.expander("🔑 API Key", expanded=False):
        api_key = st.text_input("OpenAI API key", type="password", placeholder="sk-...",
                                 help="Set OPENAI_API_KEY in Streamlit Cloud Secrets.")
        if not api_key:
            try: api_key = st.secrets["OPENAI_API_KEY"]
            except Exception: pass
        key_ok = bool(api_key and api_key.startswith("sk-"))
        if api_key and not key_ok: st.error("Key must start with sk-")
        elif key_ok: st.success("API key ready")

    # ── Navigation ────────────────────────────────────────────────────────────
    _cur_page = st.session_state.get("current_page", "Discovery")

    for section, items in _NAV_PAGES.items():
        st.markdown(f"<div class='nav-section-header'>{section}</div>", unsafe_allow_html=True)
        for page_name, icon, has_dot in items:
            is_active = (_cur_page == page_name)
            if st.button(
                f"{icon}  {page_name}",
                key=f"nav_{page_name}",
                use_container_width=True,
                type="primary" if is_active else "secondary",
            ):
                st.session_state["current_page"] = page_name
                st.rerun()

    # ── Agent controls (compact, at bottom) ──────────────────────────────────
    st.divider()
    with st.expander("🧰 Agents & Tools", expanded=False):
        s1 = st.session_state.a1_status
        s2 = st.session_state.a2_status
        st.markdown(f"""<div class="agent-card a1">
        <b>🔍 Sentinel</b> {badge(s1)}
        <small style="display:block;color:#666">{len(OS_DATA)} OS + {len(DB_DATA)} DB</small>
        </div>""", unsafe_allow_html=True)
        run_a1 = st.button("▶ Run Sentinel", width="stretch", disabled=not key_ok, key="sb_run_a1")

        st.markdown(f"""<div class="agent-card a2">
        <b>🤖 Advisor</b> {badge(s2)}
        </div>""", unsafe_allow_html=True)
        run_a2 = st.button("▶ Run Advisor", width="stretch", disabled=not key_ok, key="sb_run_a2")

        st.markdown("""<div class="agent-card a3">
        <b>🔄 Watchdog</b>
        <small style="display:block;color:#666">14-day refresh</small>
        </div>""", unsafe_allow_html=True)
        refresh_agent.render_status_card(
            st.session_state.last_refresh,
            os_count=len(st.session_state.os_df),
            db_count=len(st.session_state.db_df),
            ws_count=len(st.session_state.ws_df),
            as_count=len(st.session_state.as_df),
            fw_count=len(st.session_state.fw_df)
        )

        history = VersionGuardianAgent.get_history()
        a5s = st.session_state.get("a5_status", "idle")
        st.markdown(f"""<div class="agent-card a4">
        <b>🛡️ Guardian</b> <small>{len(history)} snapshots</small>
        </div>""", unsafe_allow_html=True)
        st.markdown(f"""<div class="agent-card a5">
        <b>🧠 Strategist</b> {badge(a5s)}
        </div>""", unsafe_allow_html=True)

        os_recs = (st.session_state.os_df["Recommendation"] != "").sum()
        db_recs = (st.session_state.db_df["Recommendation"] != "").sum()
        st.caption(f"OS: {len(st.session_state.os_df)} · DB: {len(st.session_state.db_df)} · Recs: {os_recs+db_recs}")

    with st.expander("📤 Inventory Upload", expanded=False):
        matched_os_inv, matched_db_inv = render_upload_section()
        if matched_os_inv is not None:
            with st.spinner("Matching OS inventory..."):
                enriched_os = match_os_inventory(matched_os_inv, st.session_state.os_df)
                st.session_state["inv_matched_os"] = enriched_os
        if matched_db_inv is not None:
            with st.spinner("Matching DB inventory..."):
                enriched_db = match_db_inventory(matched_db_inv, st.session_state.db_df)
                st.session_state["inv_matched_db"] = enriched_db
        stored_os_inv = st.session_state.get("inv_matched_os")
        stored_db_inv = st.session_state.get("inv_matched_db")
        if stored_os_inv is not None or stored_db_inv is not None:
            render_inventory_results(stored_os_inv, stored_db_inv)

    with st.expander("🔮 What-If Planner", expanded=False):
        render_scenario_planner(st.session_state.os_df, st.session_state.db_df)

    with st.expander("🛡️ Version History", expanded=False):
        VersionGuardianAgent.render_history_tab()


# ── Header ────────────────────────────────────────────────────────────────────
_cur_page = st.session_state.get("current_page", "Discovery")
st.markdown(f"""
<div class="infy-header">
  <h1>🖥️ INFY Migration Version Lifecycle Tracker</h1>
  <p>Infosys Enterprise Architecture &nbsp;·&nbsp;
     Powered by OpenAI GPT &nbsp;·&nbsp;
     {len(OS_DATA)} OS + {len(DB_DATA)} DB + {len(WS_DATA)} Web Servers + {len(AS_DATA)} App Servers + {len(FW_DATA)} Frameworks &nbsp;·&nbsp;
     5 AI Agents: Sentinel · Advisor · Watchdog · Guardian · Strategist &nbsp;·&nbsp;
     <strong>{_cur_page}</strong></p>
</div>
""", unsafe_allow_html=True)


# ── Agent 3 refresh banner ────────────────────────────────────────────────────
skip_until = st.session_state.get("a3_skip_until")
skip_active = skip_until and datetime.now() < skip_until

if (refresh_agent.is_refresh_due(st.session_state.last_refresh)
        and not skip_active and key_ok):
    approved = refresh_agent.render_refresh_banner(
        st.session_state.last_refresh,
        os_count=len(st.session_state.os_df),
        db_count=len(st.session_state.db_df),
        ws_count=len(st.session_state.ws_df),
        as_count=len(st.session_state.as_df),
        fw_count=len(st.session_state.fw_df)
    )
    if approved:
        VersionGuardianAgent.snapshot(
            st.session_state.os_df,
            st.session_state.db_df,
            st.session_state.changes_log,
            ws_df=st.session_state.ws_df,
            as_df=st.session_state.as_df,
            fw_df=st.session_state.fw_df
        )
        st.session_state.a1_status = "idle"
        st.session_state.a2_status = "idle"
        st.session_state.a5_os_done = False
        st.session_state.a5_db_done = False
        st.session_state.a5_ws_done = False
        st.session_state.a5_as_done = False
        st.session_state.a5_fw_done = False
        st.toast("✅ Refresh approved — click 'Run Agent 1' to check for updates.", icon="🔄")


# ── Run Agent 1 ────────────────────────────────────────────────────────────────
if run_a1:
    st.session_state.a1_status = "running"
    if not st.session_state.os_df.empty:
        VersionGuardianAgent.snapshot(
            st.session_state.os_df,
            st.session_state.db_df,
            st.session_state.changes_log,
            ws_df=st.session_state.ws_df,
            as_df=st.session_state.as_df,
            fw_df=st.session_state.fw_df
        )
    st.session_state.old_os_df = st.session_state.os_df.copy()
    st.session_state.old_db_df = st.session_state.db_df.copy()

    agent1 = OSDataAgent(api_key=api_key)

    st.info("🔍 **Agent 1** is checking the internet for lifecycle date changes. "
            "Baseline data ({} OS + {} DB) already loaded — this only looks for updates.".format(
            len(OS_DATA), len(DB_DATA)))

    prog   = st.progress(0, text="Starting internet checks...")
    status = st.empty()

    def a1_cb(pct, msg):
        prog.progress(min(pct, 1.0), text=msg)
        status.caption(msg)

    try:
        updates = agent1.fetch_updates(progress_callback=a1_cb)
        new_os, new_db, new_ws, new_as, new_fw, changes = agent1.merge_updates_into_df(
            st.session_state.os_df,
            st.session_state.db_df,
            updates,
            ws_df=st.session_state.ws_df,
            as_df=st.session_state.as_df,
            fw_df=st.session_state.fw_df
        )
        # Recompute risk scores after update
        new_os = add_risk_scores(new_os, "OS")
        new_db = add_risk_scores(new_db, "DB")
        new_ws = add_risk_scores(new_ws, "DB")
        new_as = add_risk_scores(new_as, "DB")
        new_fw = add_risk_scores(new_fw, "DB")
        st.session_state.os_df     = new_os
        st.session_state.db_df     = new_db
        st.session_state.ws_df     = new_ws
        st.session_state.as_df     = new_as
        st.session_state.fw_df     = new_fw
        st.session_state.changes_log = changes
        st.session_state.last_refresh = datetime.now()
        st.session_state.a1_status    = "done"
        st.session_state.a5_os_done   = False
        st.session_state.a5_db_done   = False
        st.session_state.a5_ws_done   = False
        st.session_state.a5_as_done   = False
        st.session_state.a5_fw_done   = False

        # ── Persist Agent 1 date changes to SQLite ────────────────────────────
        save_os_df(new_os)
        save_db_df(new_db)
        save_app_state("last_refresh", datetime.now().isoformat())

        if changes:
            st.success(f"✅ **Agent 1 complete!** Found **{len(changes)} date changes** from the internet.")
            for c in changes[:10]:
                st.markdown(f"  - {c}")
        else:
            st.success("✅ **Agent 1 complete!** No date changes found — baseline data is current.")

    except Exception as e:
        st.session_state.a1_status = "error"
        st.error(f"❌ Agent 1 error: {e}")

    st.rerun()


# ── Run Agent 2 ────────────────────────────────────────────────────────────────
if run_a2:
    st.session_state.a2_status = "running"

    # ── Checkpoint 1: Pre-flight ──────────────────────────────────────────────
    st.markdown("### 🤖 Agent 2 — Recommendation Engine")
    chk1 = st.empty()
    chk1.info("🔌 **Checkpoint 1/4** — Testing OpenAI gpt-4o-mini API connection...")

    try:
        from openai import OpenAI as _OAI2
        _c2 = _OAI2(api_key=api_key)
        _model2 = None
        for _m2 in ["gpt-4o-mini", "gpt-3.5-turbo", "gpt-3.5-turbo-0125"]:
            try:
                _r2 = _c2.chat.completions.create(
                    model=_m2, max_tokens=20,
                    messages=[{"role": "user", "content": "Reply: READY"}]
                )
                _model2 = _m2
                st.session_state["_openai_model"] = _m2
                break
            except Exception as _me2:
                if "not found" in str(_me2).lower() or "404" in str(_me2):
                    continue
                raise _me2
        if not _model2:
            raise Exception("No supported model found on this OpenAI account")
        _reply2 = _r2.choices[0].message.content.strip()
        chk1.success(f"✅ **Checkpoint 1/4 PASSED** — OpenAI `{_model2}` connected. Response: **{_reply2}**")
    except Exception as _ex2:
        chk1.error(
            f"❌ **Checkpoint 1/4 FAILED** — OpenAI API not reachable.\n\n"
            f"**Error:** `{str(_ex2)}`\n\n"
            f"Check API key and quota at platform.openai.com/usage"
        )
        st.session_state.a2_status = "error"
        st.stop()

    # ── Checkpoint 2: OS Recommendations ─────────────────────────────────────
    # Pass Guiding Principles to Agent 2 if they exist (from Strategist survey)
    _gp_data = st.session_state.get("a5_principles_table_data")
    agent2 = RecommendationAgent(api_key=api_key, guiding_principles=_gp_data)
    if _gp_data:
        st.info(f"📋 Agent 2 will align recommendations with **{len(_gp_data)} Guiding Principles** from the Strategist survey.")
    chk2 = st.empty()
    chk2.info(f"🧠 **Checkpoint 2/4** — OpenAI generating OS recommendations "
              f"({len(st.session_state.os_df)} rows, batches of 20)...")
    os_prog = st.progress(0, text="Processing OS rows...")
    os_live = st.empty()

    def a2_os_cb(pct, msg):
        os_prog.progress(min(pct, 1.0), text=msg)
        os_live.caption(f"↳ {msg}")

    try:
        new_os = agent2.generate_os_recommendations(
            st.session_state.os_df, progress_callback=a2_os_cb)
        st.session_state.os_df = new_os
        os_filled = (new_os["Recommendation"] != "").sum()
        # ── Save OS recommendations to SQLite immediately ─────────────────────
        save_os_df(new_os)
        chk2.success(f"✅ **Checkpoint 2/4 PASSED** — OS recommendations filled: "
                     f"**{os_filled} / {len(new_os)} rows** · 💾 Saved to database")
    except Exception as e:
        chk2.error(f"❌ **Checkpoint 2/4 FAILED** — OS recommendations error: `{e}`")
        st.session_state.a2_status = "error"
        st.stop()
    chk3 = st.empty()
    chk3.info(f"🧠 **Checkpoint 3/4** — OpenAI generating DB recommendations "
              f"({len(st.session_state.db_df)} rows, batches of 20)...")
    db_prog = st.progress(0, text="Processing DB rows...")
    db_live = st.empty()

    def a2_db_cb(pct, msg):
        db_prog.progress(min(pct, 1.0), text=msg)
        db_live.caption(f"↳ {msg}")

    try:
        new_db = agent2.generate_db_recommendations(
            st.session_state.db_df, progress_callback=a2_db_cb)
        st.session_state.db_df = new_db
        db_filled = (new_db["Recommendation"] != "").sum()
        # ── Save DB recommendations to SQLite immediately ─────────────────────
        save_db_df(new_db)
        chk3.success(f"✅ **Checkpoint 3/4 PASSED** — DB recommendations filled: "
                     f"**{db_filled} / {len(new_db)} rows** · 💾 Saved to database")
    except Exception as e:
        chk3.error(f"❌ **Checkpoint 3/4 FAILED** — DB recommendations error: `{e}`")
        st.session_state.a2_status = "error"
        st.stop()

    # ── Checkpoint 3.5: WS / AS / FW Recommendations ─────────────────────────
    chk35 = st.empty()
    chk35.info("🧠 **Checkpoint 3.5/4** — Generating Web Server, App Server & Framework recommendations...")
    extra_prog = st.progress(0, text="Processing WS/AS/FW...")
    extra_live = st.empty()

    try:
        def a2_extra_cb(pct, msg):
            extra_prog.progress(min(pct, 1.0), text=msg)
            extra_live.caption(f"↳ {msg}")

        new_ws = agent2.generate_generic_recommendations(
            st.session_state.ws_df, "WS", "Web Server", progress_callback=a2_extra_cb)
        st.session_state.ws_df = new_ws

        new_as = agent2.generate_generic_recommendations(
            st.session_state.as_df, "AS", "App Server", progress_callback=a2_extra_cb)
        st.session_state.as_df = new_as

        new_fw = agent2.generate_generic_recommendations(
            st.session_state.fw_df, "FW", "Framework", progress_callback=a2_extra_cb)
        st.session_state.fw_df = new_fw

        ws_f = (new_ws["Recommendation"] != "").sum()
        as_f = (new_as["Recommendation"] != "").sum()
        fw_f = (new_fw["Recommendation"] != "").sum()
        chk35.success(f"✅ **Checkpoint 3.5/4 PASSED** — WS: {ws_f}/{len(new_ws)} · "
                      f"AS: {as_f}/{len(new_as)} · FW: {fw_f}/{len(new_fw)} recommendations filled")
    except Exception as e:
        chk35.warning(f"⚠️ WS/AS/FW recommendations partial: `{e}`")

    # ── Checkpoint 4: Summary ─────────────────────────────────────────────────
    os_filled = (st.session_state.os_df["Recommendation"] != "").sum()
    db_filled = (st.session_state.db_df["Recommendation"] != "").sum()
    total     = os_filled + db_filled
    now       = datetime.now()
    save_app_state("last_refresh", now.isoformat())
    save_app_state("a2_last_run",  now.isoformat())
    st.success(
        f"✅ **Checkpoint 4/4 — Agent 2 COMPLETE** | "
        f"OS: {os_filled}/{len(st.session_state.os_df)} rows | "
        f"DB: {db_filled}/{len(st.session_state.db_df)} rows | "
        f"Total: **{total} recommendations** · 💾 **Persisted to database**"
    )

    # ── Comparison with Guiding Principles if available ─────────────────────
    _gp_data = st.session_state.get("a5_principles_table_data")
    if _gp_data:
        st.divider()
        st.subheader("📋 Agent 2 Recommendations vs Guiding Principles")
        st.caption("Cross-referencing Agent 2's technical recommendations against the Strategist's Guiding Principles.")

        compare_html = (
            "<table style='width:100%;border-collapse:collapse;font-size:0.8rem;table-layout:fixed;'>"
            "<colgroup><col style='width:10%;'/><col style='width:12%;'/>"
            "<col style='width:39%;'/><col style='width:39%;'/></colgroup>"
            "<thead><tr style='background:#1E293B;color:white;'>"
            "<th style='padding:8px;border:1px solid #334155;'>Category</th>"
            "<th style='padding:8px;border:1px solid #334155;'>Technology</th>"
            "<th style='padding:8px;border:1px solid #334155;'>🤖 Agent 2 Recommendation</th>"
            "<th style='padding:8px;border:1px solid #334155;'>📋 Guiding Principle (Strategist)</th>"
            "</tr></thead><tbody>"
        )

        # Build lookup from Agent 2 recommendations
        all_dfs = [
            (st.session_state.os_df, "OS Version", "OS"),
            (st.session_state.db_df, "Database", "Database"),
            (st.session_state.ws_df, "Web Server", "Web Server"),
            (st.session_state.as_df, "App Server", "App Server"),
            (st.session_state.fw_df, "Framework", "Framework"),
        ]
        a2_lookup = {}
        for df, name_col, cat in all_dfs:
            for _, row in df.iterrows():
                key = str(row.get(name_col, "")).lower()
                rec = str(row.get("Recommendation", ""))
                if rec.strip():
                    a2_lookup[key] = {"rec": rec, "category": cat, "name": str(row.get(name_col, ""))}

        row_idx = 0
        for gp in _gp_data:
            tech = gp.get("technology", gp.get("os_family", ""))
            cat = gp.get("category", "OS")
            gp_text = f"<strong>Upgrade:</strong> {gp.get('upgrade_principle','')}<br><strong>Replace:</strong> {gp.get('replacement_principle','')}"

            # Find matching Agent 2 recommendation
            a2_rec = ""
            for key, data in a2_lookup.items():
                if tech.lower() in key or key in tech.lower():
                    a2_rec = data["rec"]
                    break

            # Determine alignment
            if not a2_rec:
                a2_cell = "<em style='color:#9CA3AF;'>No Agent 2 recommendation yet</em>"
                align_color = "#F8FAFC"
            else:
                a2_cell = a2_rec
                # Check rough alignment
                if any(w in a2_rec.upper() for w in ["CRITICAL", "URGENT", "MIGRATE"]) and \
                   any(w in gp_text.upper() for w in ["CRITICAL", "MANDATORY", "MIGRATE"]):
                    align_color = "#FEE2E2"  # Both urgent — red tint
                elif "SUPPORTED" in a2_rec.upper() and "MONITOR" in gp_text.upper():
                    align_color = "#DCFCE7"  # Both OK — green tint
                else:
                    align_color = "#FFFBEB" if row_idx % 2 == 0 else "#FFF7ED"

            bg = align_color
            compare_html += (
                f"<tr style='background:{bg};'>"
                f"<td style='padding:6px 8px;border:1px solid #E2E8F0;font-weight:600;'>{cat}</td>"
                f"<td style='padding:6px 8px;border:1px solid #E2E8F0;font-weight:600;'>{tech}</td>"
                f"<td style='padding:6px 8px;border:1px solid #E2E8F0;color:#1E40AF;'>{a2_cell}</td>"
                f"<td style='padding:6px 8px;border:1px solid #E2E8F0;color:#065F46;'>{gp_text}</td>"
                f"</tr>"
            )
            row_idx += 1

        compare_html += "</tbody></table>"
        st.markdown(compare_html, unsafe_allow_html=True)

        aligned = sum(1 for gp in _gp_data
                      for key in a2_lookup if gp.get("technology","").lower() in key or key in gp.get("technology","").lower())
        st.caption(f"📊 {len(_gp_data)} Guiding Principles · {len(a2_lookup)} Agent 2 recommendations · {aligned} cross-referenced")

    st.session_state.a2_status = "done"
    if st.session_state.last_refresh is None:
        st.session_state.last_refresh = now

    st.rerun()


# ── Main content: Page-based routing ──────────────────────────────────────────
_show_strategist = (_cur_page == "Discovery" and st.session_state.get("show_chat", False))

# ════════════════════════════════════════════════════════════════════════════════
# DISCOVERY PAGE — Flowchart + Guiding Principles + AI Chat
# ════════════════════════════════════════════════════════════════════════════════
if _cur_page == "Discovery":
    a5s = st.session_state.get("a5_status", "idle")
    _chat_open = st.session_state.get("show_chat", False)

    # ── Flowchart using Plotly (matches hand-drawn Discovery diagram) ────────
    import plotly.graph_objects as go

    # Step ordering: maps a5_status to a numeric progress level
    # Nodes with step_ord < current level → GREEN (done)
    # Nodes with matching active_states → BLUE (active/current)
    # Nodes with step_ord > current level → GRAY (upcoming)
    _step_map_order = {
        "idle": 0, "survey": 2, "principles_table": 4,
        "chatting": 6, "principles": 7, "costing": 8,
        "ready": 9, "analysing": 9, "done": 11,
    }
    _cur_ord = _step_map_order.get(a5s, 0)

    def _fc_color(step_ord, active_states):
        if a5s in active_states:
            return "#1D4ED8"  # BLUE — currently active
        elif _cur_ord > step_ord:
            return "#065F46"  # GREEN — completed
        return "#334155"      # GRAY — upcoming

    def _fc_border(step_ord, active_states):
        if a5s in active_states:
            return "#60A5FA"
        elif _cur_ord > step_ord:
            return "#10B981"
        return "#475569"

    # Box dimensions
    BW = 1.3   # half-width
    BH = 0.6   # half-height
    GAP = 3.5  # horizontal gap between nodes

    # Node positions: (id, label, x, y, step_ord, active_states)
    # step_ord determines when this node turns green (done)
    # active_states determines when this node is blue (current)
    _nodes = [
        ("1a", "1a · Preliminary\nGuiding Principles",   0*GAP, 3, 2, ["idle","survey"]),
        ("2a", "2a · Current Landscape\nDiscovery",       1*GAP, 3, 2, ["idle","survey"]),
        ("3a", "3a · Preliminary\nDisposition",           2*GAP, 3, 4, ["principles_table"]),
        ("5",  "5 · App\nDiscussion",                     3*GAP, 3, 6, ["chatting"]),
        ("1b", "1b · Detailed\nGuiding Principles",       4*GAP, 3, 7, ["principles"]),
        ("3b", "3b · Final\nDisposition",                 5*GAP, 3, 9, ["ready","analysing"]),
        ("4b", "4b · Final\nWave Planning",               6*GAP, 3, 10, ["done"]),
        ("AP", "Approvals",                               7*GAP, 3, 11, ["done"]),
        # Upper branch
        ("4a", "4a · Preliminary\nWave Planning",         2.5*GAP, 5.5, 4, ["principles_table"]),
        ("FC", "Final Costs &\nBusiness Case",            6*GAP, 5.5, 10, ["done"]),
        # Lower branch
        ("2b", "2b · Utilization &\nCharacteristics",     1.5*GAP, 0.5, 2, ["idle","survey"]),
        ("6a", "6a · Preliminary\nCost Analysis",         2.5*GAP, 0.5, 8, ["costing"]),
    ]

    # Edges connect from box edge to box edge
    _edges = [
        # Main horizontal flow
        (0*GAP+BW, 3, 1*GAP-BW, 3),     # 1a → 2a
        (1*GAP+BW, 3, 2*GAP-BW, 3),     # 2a → 3a
        (2*GAP+BW, 3, 3*GAP-BW, 3),     # 3a → 5
        (3*GAP+BW, 3, 4*GAP-BW, 3),     # 5 → 1b
        (4*GAP+BW, 3, 5*GAP-BW, 3),     # 1b → 3b
        (5*GAP+BW, 3, 6*GAP-BW, 3),     # 3b → 4b
        (6*GAP+BW, 3, 7*GAP-BW, 3),     # 4b → Approvals
        # 4a branch UP: 3a up to 4a, 4a across to 5
        (2*GAP, 3+BH, 2.5*GAP, 5.5-BH),       # 3a ↑ 4a
        (2.5*GAP, 5.5-BH, 3*GAP, 3+BH),       # 4a ↓ 5
        # 2b/6a branch DOWN: 2a down to 2b, 2b across to 6a, 6a up to 5
        (1*GAP, 3-BH, 1.5*GAP, 0.5+BH),       # 2a ↓ 2b
        (1.5*GAP+BW, 0.5, 2.5*GAP-BW, 0.5),   # 2b → 6a
        (2.5*GAP, 0.5+BH, 3*GAP, 3-BH),       # 6a ↑ 5
        # Final Costs branch UP: 3b up to FC, FC down to 4b
        (5*GAP, 3+BH, 6*GAP, 5.5-BH),         # 3b ↑ FC
        (6*GAP, 5.5-BH, 6*GAP, 3+BH),         # FC ↓ 4b
    ]

    fig = go.Figure()

    # Separate straight (horizontal) and curved (branch) edges
    _straight_edges = _edges[:7]   # main horizontal flow
    _curved_edges = _edges[7:]     # branch connections

    # Draw straight edges with arrowheads
    for x0, y0, x1, y1 in _straight_edges:
        fig.add_annotation(
            x=x1, y=y1, ax=x0, ay=y0,
            xref="x", yref="y", axref="x", ayref="y",
            showarrow=True, arrowhead=2, arrowsize=1.5,
            arrowwidth=1.5, arrowcolor="#94A3B8",
            standoff=0, startstandoff=0
        )

    # Draw curved edges using SVG paths for branches
    import numpy as np
    for x0, y0, x1, y1 in _curved_edges:
        # Generate smooth Bezier curve points
        t = np.linspace(0, 1, 30)
        # Control points for a smooth S-curve
        cx0 = x0 + (x1 - x0) * 0.1
        cy0 = y0 + (y1 - y0) * 0.6
        cx1 = x0 + (x1 - x0) * 0.9
        cy1 = y0 + (y1 - y0) * 0.4
        # Cubic Bezier
        bx = (1-t)**3*x0 + 3*(1-t)**2*t*cx0 + 3*(1-t)*t**2*cx1 + t**3*x1
        by = (1-t)**3*y0 + 3*(1-t)**2*t*cy0 + 3*(1-t)*t**2*cy1 + t**3*y1
        fig.add_trace(go.Scatter(
            x=bx.tolist(), y=by.tolist(), mode="lines",
            line=dict(color="#94A3B8", width=1.5, dash="dot"),
            hoverinfo="skip", showlegend=False
        ))
        # Arrowhead at end
        fig.add_annotation(
            x=x1, y=y1, ax=bx[-3], ay=by[-3],
            xref="x", yref="y", axref="x", ayref="y",
            showarrow=True, arrowhead=2, arrowsize=1.2,
            arrowwidth=1.5, arrowcolor="#94A3B8",
            standoff=0, startstandoff=0
        )

    # Draw nodes
    _node_annotations = []
    for nid, label, x, y, step_ord, active_states in _nodes:
        bg = _fc_color(step_ord, active_states)
        border = _fc_border(step_ord, active_states)
        fig.add_shape(
            type="rect",
            x0=x-BW, y0=y-BH, x1=x+BW, y1=y+BH,
            fillcolor=bg, line=dict(color=border, width=2),
            layer="below"
        )
        _node_annotations.append(
            dict(x=x, y=y, text=label.replace("\n", "<br>"),
                 showarrow=False, font=dict(size=9, color="white", family="Arial"),
                 align="center")
        )

    # Layout
    fig.update_layout(
        plot_bgcolor="#0F172A",
        paper_bgcolor="#0F172A",
        xaxis=dict(visible=False, range=[-2, 7*GAP+2], fixedrange=True),
        yaxis=dict(visible=False, range=[-1, 7], scaleanchor="x", fixedrange=True),
        margin=dict(l=10, r=10, t=10, b=10),
        height=350,
        dragmode=False,
        annotations=_node_annotations + [
            dict(x=-1.5, y=6.5, text="<i>Discovery</i>", showarrow=False,
                 font=dict(size=13, color="#94A3B8"), xanchor="left"),
            dict(x=-1.5, y=-0.5, text="<i>Migrate →</i>", showarrow=False,
                 font=dict(size=12, color="#94A3B8"), xanchor="left"),
        ]
    )

    st.plotly_chart(fig, use_container_width=True, config={"displayModeBar": False})

    # ── Step-by-step guide banner ────────────────────────────────────────────
    _guide_messages = {
        "idle":             ("1", "Open AI Advisor", "Click <b>'Open AI Advisor'</b> in the sidebar to begin the Discovery questionnaire.", "#3B82F6", "👆"),
        "survey":           ("2", "Complete Survey", "Fill in all 4 category tabs (Cloud, Upgrade, Replacement, Principles) then click <b>'Confirm All Categories'</b>.", "#7C3AED", "📝"),
        "principles_table": ("3", "Review Principles", "Review the generated Guiding Principles, costs, and wave plan. Then click <b>'Proceed to Policy Chat'</b>.", "#059669", "📋"),
        "chatting":         ("4", "Policy Chat", "Answer Agent 5's questions about compliance, budgets, and risk tolerance. After 4+ exchanges, click <b>'Proceed to Guiding Principles'</b>.", "#D97706", "💬"),
        "principles":       ("5", "Generating GP", "Agent 5 is synthesising your answers into Guiding Principles. Please wait...", "#6366F1", "⚙️"),
        "costing":          ("5", "Cost Intel", "Agent 5 is fetching live vendor pricing. Please wait...", "#6366F1", "⚙️"),
        "ready":            ("6", "Ready to Analyse", "Review the principles and costs, then click <b>'Generate Final Recommendations'</b>.", "#DC2626", "🚀"),
        "analysing":        ("6", "Analysing", "Agent 5 is generating Final Recommendations across all categories. Please wait...", "#6366F1", "⚙️"),
        "done":             ("7", "Complete!", "Final Recommendations added to all tabs. <b>Download Excel/PDF</b> below.", "#059669", "✅"),
    }
    _gm = _guide_messages.get(a5s, _guide_messages["idle"])
    _step_num, _step_title, _step_msg, _step_color, _step_icon = _gm

    st.markdown(f"""
    <div style="display:flex;align-items:center;gap:14px;
                background:linear-gradient(135deg,{_step_color}15,{_step_color}08);
                border:1px solid {_step_color}40;border-left:4px solid {_step_color};
                border-radius:0 10px 10px 0;padding:0.8rem 1.2rem;margin:0.5rem 0 1rem;">
      <div style="font-size:1.8rem;">{_step_icon}</div>
      <div>
        <div style="font-size:0.7rem;color:{_step_color};font-weight:700;letter-spacing:1px;text-transform:uppercase;">
          Step {_step_num} — {_step_title}
        </div>
        <div style="font-size:0.88rem;color:#334155;margin-top:2px;">
          {_step_msg}
        </div>
      </div>
    </div>
    """, unsafe_allow_html=True)

    # ── Guiding Principles — 4 Category Cards ────────────────────────────────
    st.markdown("""
    <div style="text-align:center;margin-bottom:1rem;">
        <h2 style="color:#1E293B;margin:0;">Guiding Principles</h2>
        <p style="color:#64748B;font-size:0.9rem;margin:4px 0 0;">
            Define your migration strategy across 5 technology categories
        </p>
    </div>
    """, unsafe_allow_html=True)

    gp_cols = st.columns(4)
    _gp_categories = [
        ("☁️", "Cat 1:\nPreferred Cloud", "Cloud target selection\n& OS-to-cloud mapping", "#1E3A8A"),
        ("⬆️", "Cat 2:\nUpgrade Mandates", "OS & DB upgrade paths\nversion lifecycle rules", "#065F46"),
        ("🔁", "Cat 2:\nReplacement Mandates", "Web/App/Framework\nEOL replacement rules", "#991B1B"),
        ("📋", "Cat 3:\nPrinciples & Costs", "Interactive policy,\ncompliance, budgets", "#92400E"),
    ]
    for col, (icon, title, desc, color) in zip(gp_cols, _gp_categories):
        with col:
            title_html = title.replace("\n", "<br>")
            desc_html = desc.replace("\n", "<br>")
            st.markdown(f"""
            <div class="gp-card" style="border-top:4px solid {color};">
                <div class="gp-icon">{icon}</div>
                <div class="gp-title">{title_html}</div>
                <div class="gp-desc">{desc_html}</div>
            </div>
            """, unsafe_allow_html=True)

    # ── Flashing indicator bottom-right (visual, points to sidebar) ─────────
    if not _chat_open:
        st.markdown("""
        <style>
        .ai-float-badge {
            position: fixed; bottom: 24px; right: 24px; z-index: 9999;
            background: linear-gradient(135deg,#7C3AED,#9333EA);
            color: white; border-radius: 50px; padding: 12px 20px;
            font-size: 0.85rem; font-weight: 700;
            display: flex; align-items: center; gap: 8px;
            box-shadow: 0 4px 20px rgba(124,58,237,0.5);
            animation: ai-pulse 2s ease-in-out infinite;
            cursor: default;
        }
        @keyframes ai-pulse {
            0%, 100% { box-shadow: 0 4px 20px rgba(124,58,237,0.5); transform: scale(1); }
            50% { box-shadow: 0 4px 30px rgba(124,58,237,0.9), 0 0 50px rgba(124,58,237,0.2); transform: scale(1.03); }
        }
        .ai-dot { width: 10px; height: 10px; border-radius: 50%;
                   background: #34D399; animation: ai-dot-blink 1s infinite; }
        @keyframes ai-dot-blink { 0%,100%{opacity:1;} 50%{opacity:0.2;} }
        </style>
        <div class="ai-float-badge">
            <div class="ai-dot"></div>
            🧠 AI Advisor ← Sidebar
        </div>
        """, unsafe_allow_html=True)

    if not _chat_open:
        _show_strategist = False


# ════════════════════════════════════════════════════════════════════════════════
# VERSION LIFECYCLE PAGE — Original tabs
# ════════════════════════════════════════════════════════════════════════════════
if _cur_page == "Version Lifecycle":
    tab_dash, tab_os, tab_db, tab_wsas, tab_fw = st.tabs([
        "📊 Dashboard", "🖥️ OS Versions", "🗄️ DB Versions",
        "🌐 Web & App Servers", "📦 Frameworks",
    ])
elif _cur_page != "Discovery":
    # Placeholder pages for other navigation items
    st.markdown(f"""
    <div style="text-align:center;padding:4rem 2rem;">
        <div style="font-size:3rem;margin-bottom:1rem;">🚧</div>
        <h2 style="color:#1E293B;">{_cur_page}</h2>
        <p style="color:#64748B;font-size:1rem;">
            This module is coming soon. Currently, <strong>Discovery</strong> and
            <strong>Version Lifecycle</strong> are fully operational.
        </p>
    </div>
    """, unsafe_allow_html=True)

# When NOT on Version Lifecycle, we still need tab variables for the existing code.
# Create hidden tabs so the `with tab_dash:` blocks don't error.
if _cur_page != "Version Lifecycle":
    tab_dash, tab_os, tab_db, tab_wsas, tab_fw = st.tabs([
        "📊 Dashboard", "🖥️ OS Versions", "🗄️ DB Versions",
        "🌐 Web & App Servers", "📦 Frameworks",
    ])
    # Hide these tabs with CSS — they exist only to avoid NameError
    st.markdown("<style>.stTabs{display:none!important;}</style>", unsafe_allow_html=True)


# ────────────────── Tab 0: Executive Dashboard ────────────────────────────────
with tab_dash:
    st.subheader("📊 Executive Risk Dashboard")
    st.caption("Real-time risk analysis across your entire OS, DB, Web Server, App Server & Framework portfolio")

    # Top-level KPI metrics
    os_df_d = st.session_state.os_df
    db_df_d = st.session_state.db_df
    risk_summary = get_risk_summary(os_df_d, db_df_d)
    total_items = len(os_df_d) + len(db_df_d) + len(st.session_state.ws_df) + len(st.session_state.as_df) + len(st.session_state.fw_df)

    k1, k2, k3, k4, k5, k6 = st.columns(6)
    with k1: st.metric("Total Portfolio", total_items)
    with k2: st.metric("🔴 Critical", risk_summary["CRITICAL"])
    with k3: st.metric("🟠 High", risk_summary["HIGH"])
    with k4: st.metric("🟡 Medium", risk_summary["MEDIUM"])
    with k5: st.metric("🟢 Low", risk_summary["LOW"])
    with k6: st.metric("✅ Minimal", risk_summary["MINIMAL"])

    st.divider()

    # Charts row 1
    ch1, ch2 = st.columns(2)
    with ch1:
        fig_risk = risk_distribution_chart(os_df_d, db_df_d)
        st.plotly_chart(fig_risk, use_container_width=True)
    with ch2:
        fig_bar = status_breakdown_chart(os_df_d, db_df_d)
        st.plotly_chart(fig_bar, use_container_width=True)

    st.divider()

    # EOL Timeline
    fig_timeline = eol_timeline_chart(os_df_d, db_df_d)
    st.plotly_chart(fig_timeline, use_container_width=True)

    st.divider()

    # Top urgent items table
    ch3, ch4 = st.columns([3, 2])
    with ch3:
        st.subheader("🚨 Top 10 Most Urgent Items")
        urgent = top_urgent_items(os_df_d, db_df_d, n=10)
        if not urgent.empty:
            st.dataframe(urgent, height=380, hide_index=True, use_container_width=True)
        else:
            st.info("Risk scores not yet computed.")
    with ch4:
        fig_hist = risk_score_histogram(os_df_d, db_df_d)
        st.plotly_chart(fig_hist, use_container_width=True)

    # Quick filter: items expiring during project window
    st.divider()
    st.subheader("⚠️ Items Expiring During Project Window")
    from datetime import date as _date
    proj_end = get_project_end()
    proj_start = st.session_state.get("project_start", _date(2026, 4, 1))
    expiring_os = os_df_d[
        os_df_d["Days Until EOL"].apply(
            lambda d: d is not None and 0 < d <= (proj_end - _date.today()).days
        )
    ] if "Days Until EOL" in os_df_d.columns else pd.DataFrame()
    expiring_db = db_df_d[
        db_df_d["Days Until EOL"].apply(
            lambda d: d is not None and 0 < d <= (proj_end - _date.today()).days
        )
    ] if "Days Until EOL" in db_df_d.columns else pd.DataFrame()

    exp_count = len(expiring_os) + len(expiring_db)
    if exp_count > 0:
        st.warning(f"**{exp_count} items** have support ending before {proj_end.strftime('%d %b %Y')}")
        ex1, ex2 = st.columns(2)
        with ex1:
            if not expiring_os.empty:
                st.caption(f"OS: {len(expiring_os)} items expiring")
                show_cols = [c for c in ["OS Version", "Days Until EOL", "Risk Score", "Risk Level", "Recommendation"] if c in expiring_os.columns]
                st.dataframe(expiring_os[show_cols].sort_values("Days Until EOL"), hide_index=True, height=250)
        with ex2:
            if not expiring_db.empty:
                st.caption(f"DB: {len(expiring_db)} items expiring")
                show_cols = [c for c in ["Database", "Version", "Days Until EOL", "Risk Score", "Risk Level", "Recommendation"] if c in expiring_db.columns]
                st.dataframe(expiring_db[show_cols].sort_values("Days Until EOL"), hide_index=True, height=250)
    else:
        st.success("No items expiring during the project window.")


# ────────────────── Tab 1: OS Versions ────────────────────────────────────────
with tab_os:
    os_df = st.session_state.os_df

    m1, m2, m3, m4, m5 = st.columns(5)

    def is_eol(row):
        for col in ["Extended/LTSC Support End", "Mainstream/Full Support End"]:
            v = str(row.get(col, ""))
            if any(k in v.lower() for k in ["ended", "end of"]):
                return True
            if len(v) == 10 and v[0].isdigit():
                try:
                    from datetime import datetime as dt
                    return dt.strptime(v, "%Y-%m-%d") < dt.now()
                except Exception:
                    pass
        return False

    eol_count  = sum(1 for _, r in os_df.iterrows() if is_eol(r))
    upg_count  = (os_df.get("Upgrade",  pd.Series(dtype=str)) == "Y").sum()
    repl_count = (os_df.get("Replace",  pd.Series(dtype=str)) == "Y").sum()
    rec_count  = (os_df["Recommendation"] != "").sum()

    with m1: st.metric("Total OS Entries",   len(os_df))
    with m2: st.metric("⚠️ EOL / Expired",   int(eol_count))
    with m3: st.metric("⬆️ Upgrade Flagged", int(upg_count))
    with m4: st.metric("🔁 Replace Flagged", int(repl_count))
    with m5: st.metric("💡 Recommendations", int(rec_count))

    with st.expander("🔍 Filters", expanded=False):
        fc1, fc2, fc3, fc4 = st.columns(4)
        with fc1:
            families = ["All"] + sorted(set(
                v.split()[0] for v in os_df["OS Version"].dropna() if v and not v.startswith("[")
            ))
            fam_f = st.selectbox("OS Family", families)
        with fc2: upg_f  = st.selectbox("Upgrade", ["All","Y","N"])
        with fc3: repl_f = st.selectbox("Replace", ["All","Y","N"])
        with fc4: rec_f  = st.selectbox("Has Recommendation", ["All","Yes","No"])

    view = os_df.copy()
    # Global search filter
    gs = st.session_state.get("global_search", "").strip()
    if gs:
        view = view[view.apply(lambda r: gs.lower() in " ".join(str(v) for v in r.values).lower(), axis=1)]
    if fam_f  != "All": view = view[view["OS Version"].str.startswith(fam_f, na=False)]
    if upg_f  != "All": view = view[view.get("Upgrade", pd.Series(dtype=str)) == upg_f]
    if repl_f != "All": view = view[view.get("Replace", pd.Series(dtype=str)) == repl_f]
    if rec_f  == "Yes": view = view[view["Recommendation"] != ""]
    if rec_f  == "No":  view = view[view["Recommendation"] == ""]

    disp = {
        "OS Version":                    st.column_config.TextColumn("OS Version",        width=220),
        "Availability Date":             st.column_config.TextColumn("Available",         width=110),
        "Security/Standard Support End": st.column_config.TextColumn("Security End",      width=120),
        "Mainstream/Full Support End":   st.column_config.TextColumn("Mainstream End",    width=140),
        "Extended/LTSC Support End":     st.column_config.TextColumn("Extended End",      width=120),
        "Min CPU":                       st.column_config.TextColumn("Min CPU",           width=140),
        "Min RAM":                       st.column_config.TextColumn("Min RAM",           width=80),
        "Notes":                         st.column_config.TextColumn("Notes",             width=160),
        "Recommendation":                st.column_config.TextColumn("💡 Recommendation", width=380),
        "Upgrade":                       st.column_config.TextColumn("⬆ Upgrade",         width=80),
        "Replace":                       st.column_config.TextColumn("🔁 Replace",        width=80),
        "Primary Alternative":           st.column_config.TextColumn("Primary Alt",       width=160),
    }
    if "Risk Score" in view.columns:
        disp["Risk Score"] = st.column_config.NumberColumn("🎯 Risk", width=70)
    if "Risk Level" in view.columns:
        disp["Risk Level"] = st.column_config.TextColumn("Risk Level", width=90)
    if "Days Until EOL" in view.columns:
        disp["Days Until EOL"] = st.column_config.NumberColumn("📅 Days to EOL", width=100)
    if "Policy Recommendation" in view.columns:
        disp["Policy Recommendation"] = st.column_config.TextColumn("🏛️ Policy Rec", width=360)
    if "Verdict" in view.columns:
        disp["Verdict"] = st.column_config.TextColumn("Verdict", width=120)
    # Hide internal color column
    hide_cols = [c for c in ["Risk Color"] if c in view.columns]
    if hide_cols:
        view = view.drop(columns=hide_cols)

    st.caption(f"Showing {len(view)} of {len(os_df)} OS entries")
    st.dataframe(view, width="stretch", height=520, hide_index=True, column_config=disp)


# ────────────────── Tab 2: DB Versions ────────────────────────────────────────
with tab_db:
    db_df = st.session_state.db_df

    dm1, dm2, dm3, dm4, dm5 = st.columns(5)
    eol_db = (db_df.get("Status", pd.Series(dtype=str)).str.lower() == "end of life").sum()
    exp_db = (db_df.get("Status", pd.Series(dtype=str)).str.lower() == "expiring soon").sum()
    sup_db = (db_df.get("Status", pd.Series(dtype=str)).str.lower() == "supported").sum()
    rec_db = (db_df["Recommendation"] != "").sum()

    with dm1: st.metric("Total DB Entries",   len(db_df))
    with dm2: st.metric("🔴 End of Life",     int(eol_db))
    with dm3: st.metric("🟡 Expiring Soon",   int(exp_db))
    with dm4: st.metric("🟢 Supported",       int(sup_db))
    with dm5: st.metric("💡 Recommendations", int(rec_db))

    with st.expander("🔍 Filters", expanded=False):
        df1, df2, df3, df4 = st.columns(4)
        with df1:
            dbs = ["All"] + sorted(db_df["Database"].dropna().unique().tolist())
            db_f = st.selectbox("Database", dbs)
        with df2:
            types_ = ["All"] + sorted(db_df["Type"].dropna().unique().tolist())
            type_f = st.selectbox("Type", types_)
        with df3:
            statuses = ["All"] + sorted(db_df["Status"].dropna().unique().tolist())
            stat_f = st.selectbox("Status", statuses)
        with df4:
            repl_db_f = st.selectbox("Replace", ["All","Y","N"], key="repl_db")

    view_db = db_df.copy()
    # Global search filter
    gs = st.session_state.get("global_search", "").strip()
    if gs:
        view_db = view_db[view_db.apply(lambda r: gs.lower() in " ".join(str(v) for v in r.values).lower(), axis=1)]
    if db_f      != "All": view_db = view_db[view_db["Database"] == db_f]
    if type_f    != "All": view_db = view_db[view_db["Type"]     == type_f]
    if stat_f    != "All": view_db = view_db[view_db["Status"]   == stat_f]
    if repl_db_f != "All": view_db = view_db[view_db.get("Replace", pd.Series(dtype=str)) == repl_db_f]

    def _style_status(val):
        m = {"end of life":   "background-color:#FFCDD2;color:#B71C1C",
             "expiring soon": "background-color:#FFE0B2;color:#E65100",
             "supported":     "background-color:#C8E6C9;color:#1B5E20",
             "future":        "background-color:#E3F2FD;color:#0D47A1"}
        return m.get(str(val).lower(), "")

    db_disp = {
        "Database":                st.column_config.TextColumn("Database",        width=180),
        "Version":                 st.column_config.TextColumn("Version",         width=120),
        "Type":                    st.column_config.TextColumn("Type",            width=150),
        "Mainstream / Premier End":st.column_config.TextColumn("Mainstream End",  width=130),
        "Extended Support End":    st.column_config.TextColumn("Extended End",    width=120),
        "Status":                  st.column_config.TextColumn("Status",          width=120),
        "Min CPU":                 st.column_config.TextColumn("Min CPU",         width=120),
        "Min RAM":                 st.column_config.TextColumn("Min RAM",         width=80),
        "Notes":                   st.column_config.TextColumn("Notes",           width=180),
        "Recommendation":          st.column_config.TextColumn("💡 Recommendation",width=360),
        "Upgrade":                 st.column_config.TextColumn("⬆ Upgrade",       width=80),
        "Replace":                 st.column_config.TextColumn("🔁 Replace",      width=80),
        "Primary Alternative":     st.column_config.TextColumn("Primary Alt",     width=150),
    }
    if "Risk Score" in view_db.columns:
        db_disp["Risk Score"] = st.column_config.NumberColumn("🎯 Risk", width=70)
    if "Risk Level" in view_db.columns:
        db_disp["Risk Level"] = st.column_config.TextColumn("Risk Level", width=90)
    if "Days Until EOL" in view_db.columns:
        db_disp["Days Until EOL"] = st.column_config.NumberColumn("📅 Days to EOL", width=100)
    if "Policy Recommendation" in view_db.columns:
        db_disp["Policy Recommendation"] = st.column_config.TextColumn("🏛️ Policy Rec", width=360)
    if "Verdict" in view_db.columns:
        db_disp["Verdict"] = st.column_config.TextColumn("Verdict", width=120)
    # Hide internal color column
    hide_cols = [c for c in ["Risk Color"] if c in view_db.columns]
    if hide_cols:
        view_db = view_db.drop(columns=hide_cols)

    st.caption(f"Showing {len(view_db)} of {len(db_df)} DB entries")
    st.dataframe(
        view_db.style.map(_style_status, subset=["Status"]) if "Status" in view_db.columns else view_db,
        width="stretch", height=520, hide_index=True, column_config=db_disp
    )


# ────────────────── Tab: Web & App Servers ─────────────────────────────────────
with tab_wsas:
    ws_df = st.session_state.ws_df
    st.subheader("🌐 Web Server Versions")
    st.caption("Lifecycle tracking for IIS, Nginx, Apache and other web servers")

    wm1, wm2, wm3, wm4 = st.columns(4)
    eol_ws = (ws_df.get("Status", pd.Series(dtype=str)).str.lower() == "end of life").sum()
    exp_ws = (ws_df.get("Status", pd.Series(dtype=str)).str.lower() == "expiring soon").sum()
    sup_ws = (ws_df.get("Status", pd.Series(dtype=str)).str.lower() == "supported").sum()
    with wm1: st.metric("Total Web Servers", len(ws_df))
    with wm2: st.metric("🔴 End of Life", int(eol_ws))
    with wm3: st.metric("🟡 Expiring Soon", int(exp_ws))
    with wm4: st.metric("🟢 Supported", int(sup_ws))

    with st.expander("🔍 Filters", expanded=False):
        wf1, wf2, wf3 = st.columns(3)
        with wf1:
            ws_names = ["All"] + sorted(ws_df["Web Server"].dropna().unique().tolist())
            ws_f = st.selectbox("Web Server", ws_names, key="ws_filter")
        with wf2:
            ws_statuses = ["All"] + sorted(ws_df["Status"].dropna().unique().tolist())
            ws_stat_f = st.selectbox("Status", ws_statuses, key="ws_stat_filter")
        with wf3:
            ws_upg_f = st.selectbox("Upgrade", ["All","Y","N"], key="ws_upg_filter")

    view_ws = ws_df.copy()
    if ws_f != "All": view_ws = view_ws[view_ws["Web Server"] == ws_f]
    if ws_stat_f != "All": view_ws = view_ws[view_ws["Status"] == ws_stat_f]
    if ws_upg_f != "All": view_ws = view_ws[view_ws.get("Upgrade", pd.Series(dtype=str)) == ws_upg_f]

    def _style_ws_status(val):
        m = {"end of life": "background-color:#FFCDD2;color:#B71C1C",
             "expiring soon": "background-color:#FFE0B2;color:#E65100",
             "supported": "background-color:#C8E6C9;color:#1B5E20"}
        return m.get(str(val).lower(), "")

    ws_disp = {
        "Web Server": st.column_config.TextColumn("Web Server", width=140),
        "Version": st.column_config.TextColumn("Version", width=100),
        "Type": st.column_config.TextColumn("Type", width=120),
        "Mainstream / Premier End": st.column_config.TextColumn("Mainstream End", width=130),
        "Extended Support End": st.column_config.TextColumn("Extended End", width=120),
        "Status": st.column_config.TextColumn("Status", width=120),
        "Notes": st.column_config.TextColumn("Notes", width=220),
        "Recommendation": st.column_config.TextColumn("💡 Recommendation", width=360),
        "Upgrade": st.column_config.TextColumn("⬆ Upgrade", width=80),
        "Replace": st.column_config.TextColumn("🔁 Replace", width=80),
        "Primary Alternative": st.column_config.TextColumn("Primary Alt", width=150),
    }
    if "Risk Score" in view_ws.columns:
        ws_disp["Risk Score"] = st.column_config.NumberColumn("🎯 Risk", width=70)
    if "Risk Level" in view_ws.columns:
        ws_disp["Risk Level"] = st.column_config.TextColumn("Risk Level", width=90)
    if "Days Until EOL" in view_ws.columns:
        ws_disp["Days Until EOL"] = st.column_config.NumberColumn("📅 Days to EOL", width=100)
    hide_cols = [c for c in ["Risk Color"] if c in view_ws.columns]
    if hide_cols: view_ws = view_ws.drop(columns=hide_cols)

    st.caption(f"Showing {len(view_ws)} of {len(ws_df)} Web Server entries")
    st.dataframe(
        view_ws.style.map(_style_ws_status, subset=["Status"]) if "Status" in view_ws.columns else view_ws,
        width="stretch", height=520, hide_index=True, column_config=ws_disp
    )


    # ────────────────── Section: Application Servers ─────────────────────────
    st.divider()
    as_df = st.session_state.as_df
    st.subheader("⚙️ Application Server Versions")
    st.caption("Lifecycle tracking for Tomcat, JBoss, WebSphere, Kafka, RabbitMQ, Kubernetes and more")

    am1, am2, am3, am4 = st.columns(4)
    eol_as = (as_df.get("Status", pd.Series(dtype=str)).str.lower() == "end of life").sum()
    exp_as = (as_df.get("Status", pd.Series(dtype=str)).str.lower() == "expiring soon").sum()
    sup_as = (as_df.get("Status", pd.Series(dtype=str)).str.lower() == "supported").sum()
    with am1: st.metric("Total App Servers", len(as_df))
    with am2: st.metric("🔴 End of Life", int(eol_as))
    with am3: st.metric("🟡 Expiring Soon", int(exp_as))
    with am4: st.metric("🟢 Supported", int(sup_as))

    with st.expander("🔍 Filters", expanded=False):
        af1, af2, af3 = st.columns(3)
        with af1:
            as_names = ["All"] + sorted(as_df["App Server"].dropna().unique().tolist())
            as_f = st.selectbox("App Server", as_names, key="as_filter")
        with af2:
            as_types = ["All"] + sorted(as_df["Type"].dropna().unique().tolist())
            as_type_f = st.selectbox("Type", as_types, key="as_type_filter")
        with af3:
            as_statuses = ["All"] + sorted(as_df["Status"].dropna().unique().tolist())
            as_stat_f = st.selectbox("Status", as_statuses, key="as_stat_filter")

    view_as = as_df.copy()
    if as_f != "All": view_as = view_as[view_as["App Server"] == as_f]
    if as_type_f != "All": view_as = view_as[view_as["Type"] == as_type_f]
    if as_stat_f != "All": view_as = view_as[view_as["Status"] == as_stat_f]

    def _style_as_status(val):
        m = {"end of life": "background-color:#FFCDD2;color:#B71C1C",
             "expiring soon": "background-color:#FFE0B2;color:#E65100",
             "supported": "background-color:#C8E6C9;color:#1B5E20"}
        return m.get(str(val).lower(), "")

    as_disp = {
        "App Server": st.column_config.TextColumn("App Server", width=160),
        "Version": st.column_config.TextColumn("Version", width=100),
        "Type": st.column_config.TextColumn("Type", width=160),
        "Mainstream / Premier End": st.column_config.TextColumn("Mainstream End", width=130),
        "Extended Support End": st.column_config.TextColumn("Extended End", width=120),
        "Status": st.column_config.TextColumn("Status", width=120),
        "Notes": st.column_config.TextColumn("Notes", width=220),
        "Recommendation": st.column_config.TextColumn("💡 Recommendation", width=360),
        "Upgrade": st.column_config.TextColumn("⬆ Upgrade", width=80),
        "Replace": st.column_config.TextColumn("🔁 Replace", width=80),
        "Primary Alternative": st.column_config.TextColumn("Primary Alt", width=150),
    }
    if "Risk Score" in view_as.columns:
        as_disp["Risk Score"] = st.column_config.NumberColumn("🎯 Risk", width=70)
    if "Risk Level" in view_as.columns:
        as_disp["Risk Level"] = st.column_config.TextColumn("Risk Level", width=90)
    if "Days Until EOL" in view_as.columns:
        as_disp["Days Until EOL"] = st.column_config.NumberColumn("📅 Days to EOL", width=100)
    hide_cols = [c for c in ["Risk Color"] if c in view_as.columns]
    if hide_cols: view_as = view_as.drop(columns=hide_cols)

    st.caption(f"Showing {len(view_as)} of {len(as_df)} App Server entries")
    st.dataframe(
        view_as.style.map(_style_as_status, subset=["Status"]) if "Status" in view_as.columns else view_as,
        width="stretch", height=520, hide_index=True, column_config=as_disp
    )


# ────────────────── Tab: Frameworks ───────────────────────────────────────────
with tab_fw:
    fw_df = st.session_state.fw_df
    st.subheader("📦 Framework & Runtime Versions")
    st.caption("Lifecycle tracking for .NET, React, Angular, Spring Boot, Node.js, Java, Python, PHP and more")

    fm1, fm2, fm3, fm4 = st.columns(4)
    eol_fw = (fw_df.get("Status", pd.Series(dtype=str)).str.lower() == "end of life").sum()
    exp_fw = (fw_df.get("Status", pd.Series(dtype=str)).str.lower() == "expiring soon").sum()
    sup_fw = (fw_df.get("Status", pd.Series(dtype=str)).str.lower() == "supported").sum()
    with fm1: st.metric("Total Frameworks", len(fw_df))
    with fm2: st.metric("🔴 End of Life", int(eol_fw))
    with fm3: st.metric("🟡 Expiring Soon", int(exp_fw))
    with fm4: st.metric("🟢 Supported", int(sup_fw))

    with st.expander("🔍 Filters", expanded=False):
        ff1, ff2, ff3 = st.columns(3)
        with ff1:
            fw_names = ["All"] + sorted(fw_df["Framework"].dropna().unique().tolist())
            fw_f = st.selectbox("Framework", fw_names, key="fw_filter")
        with ff2:
            fw_types = ["All"] + sorted(fw_df["Type"].dropna().unique().tolist())
            fw_type_f = st.selectbox("Type", fw_types, key="fw_type_filter")
        with ff3:
            fw_statuses = ["All"] + sorted(fw_df["Status"].dropna().unique().tolist())
            fw_stat_f = st.selectbox("Status", fw_statuses, key="fw_stat_filter")

    view_fw = fw_df.copy()
    if fw_f != "All": view_fw = view_fw[view_fw["Framework"] == fw_f]
    if fw_type_f != "All": view_fw = view_fw[view_fw["Type"] == fw_type_f]
    if fw_stat_f != "All": view_fw = view_fw[view_fw["Status"] == fw_stat_f]

    def _style_fw_status(val):
        m = {"end of life": "background-color:#FFCDD2;color:#B71C1C",
             "expiring soon": "background-color:#FFE0B2;color:#E65100",
             "supported": "background-color:#C8E6C9;color:#1B5E20",
             "future": "background-color:#E3F2FD;color:#0D47A1"}
        return m.get(str(val).lower(), "")

    fw_disp = {
        "Framework": st.column_config.TextColumn("Framework", width=160),
        "Version": st.column_config.TextColumn("Version", width=120),
        "Type": st.column_config.TextColumn("Type", width=160),
        "Mainstream / Premier End": st.column_config.TextColumn("Mainstream End", width=130),
        "Extended Support End": st.column_config.TextColumn("Extended End", width=120),
        "Status": st.column_config.TextColumn("Status", width=120),
        "Notes": st.column_config.TextColumn("Notes", width=220),
        "Recommendation": st.column_config.TextColumn("💡 Recommendation", width=360),
        "Upgrade": st.column_config.TextColumn("⬆ Upgrade", width=80),
        "Replace": st.column_config.TextColumn("🔁 Replace", width=80),
        "Primary Alternative": st.column_config.TextColumn("Primary Alt", width=150),
    }
    if "Risk Score" in view_fw.columns:
        fw_disp["Risk Score"] = st.column_config.NumberColumn("🎯 Risk", width=70)
    if "Risk Level" in view_fw.columns:
        fw_disp["Risk Level"] = st.column_config.TextColumn("Risk Level", width=90)
    if "Days Until EOL" in view_fw.columns:
        fw_disp["Days Until EOL"] = st.column_config.NumberColumn("📅 Days to EOL", width=100)
    hide_cols = [c for c in ["Risk Color"] if c in view_fw.columns]
    if hide_cols: view_fw = view_fw.drop(columns=hide_cols)

    st.caption(f"Showing {len(view_fw)} of {len(fw_df)} Framework entries")
    st.dataframe(
        view_fw.style.map(_style_fw_status, subset=["Status"]) if "Status" in view_fw.columns else view_fw,
        width="stretch", height=520, hide_index=True, column_config=fw_disp
    )


# ────────────────── Strategist Panel (shown on Discovery page) ─────────────────
if _show_strategist:
    if not key_ok:
        st.warning("Enter your OpenAI API key in the sidebar (API Key expander) to use Agent 5.")
    else:
        a5s = st.session_state.get("a5_status", "idle")

        # ── PHASE 0: DISCOVERY SURVEY (PDF Categories 1-3) ─────────────────────
        if a5s in ("idle", "survey"):
            from agents.agent_analysis import categorize_os_families, get_family_display
            from agents.agent_analysis import categorize_db_families, get_db_family_display

            if a5s == "idle":
                st.session_state.a5_status = "survey"
                st.rerun()

            # Categorize families from live data
            os_families = categorize_os_families(st.session_state.os_df)
            st.session_state.a5_landscape_families = os_families
            db_families = categorize_db_families(st.session_state.db_df)
            st.session_state.a5_db_landscape_families = db_families
            family_display = get_family_display()
            db_family_display = get_db_family_display()

            ws_df = st.session_state.ws_df
            as_df = st.session_state.as_df
            fw_df = st.session_state.fw_df
            ws_names = sorted(ws_df["Web Server"].dropna().unique().tolist())
            as_names = sorted(as_df["App Server"].dropna().unique().tolist())
            fw_names = sorted(fw_df["Framework"].dropna().unique().tolist())

            st.caption("Complete the categories below. Agent 1's baseline data is pre-loaded — confirm what's in your environment.")

            # ── 4 Category tabs matching the Guiding Principles PDF ──────────
            _t_cloud, _t_upgrade, _t_replace, _t_policy = st.tabs([
                "☁️ Cat 1: Preferred Cloud",
                "⬆️ Cat 2: Upgrade Mandates",
                "🔁 Cat 2: Replacement Mandates",
                "📋 Cat 3: Principles & Costs",
            ])

            # ══════════════════════════════════════════════════════════════════
            # CATEGORY 1: PREFERRED CLOUD PROVIDER
            # ══════════════════════════════════════════════════════════════════
            with _t_cloud:
                st.markdown("""
                <div style="background:#EFF6FF;border:1px solid #BFDBFE;border-radius:8px;
                            padding:0.8rem 1rem;margin-bottom:0.8rem;">
                  <h4 style="margin:0;color:#1E3A8A;">Category 1: Preferred Cloud Provider</h4>
                  <small style="color:#64748B;">Select the cloud profile that best matches your target.
                  This determines upgrade paths and replacement targets for all technologies.</small>
                </div>""", unsafe_allow_html=True)

                CLOUD_PROFILES = [
                    ("Microsoft-Centric (Azure)", "🔷",
                     "Windows Server (2012-2025), Active Directory, Azure Hybrid Benefit, SQL Server on Azure",
                     "Azure"),
                    ("Linux & Scale (AWS)", "🟠",
                     "RHEL/RPM-based footprints, Graviton ARM, 40% better price-performance on modern Linux",
                     "AWS"),
                    ("Container & Data (GCP)", "🔵",
                     "Ubuntu/Debian, Kubernetes (GKE), AI data pipelines, global private fiber",
                     "GCP"),
                    ("Database & Legacy Bridge (Oracle/OCI)", "🔴",
                     "Oracle Linux, Solaris, Oracle Database backends, most stable legacy path",
                     "OCI"),
                    ("Sovereign-First (US Gov / High-Reg)", "🛡️",
                     "FedRAMP High, ITAR, CJIS — vendor-agnostic, compliance-driven",
                     "GovCloud"),
                ]
                custom_profiles = st.session_state.get("a5_custom_cloud_profiles", [])
                all_profiles = CLOUD_PROFILES + custom_profiles

                # Cloud-to-OS mapping table
                st.markdown("""
                <table style='width:100%;border-collapse:collapse;font-size:0.82rem;margin-bottom:1rem;'>
                <thead><tr style='background:#1E3A8A;color:white;'>
                <th style='padding:8px;border:1px solid #334155;'>OS Family</th>
                <th style='padding:8px;border:1px solid #334155;'>Workload Type</th>
                <th style='padding:8px;border:1px solid #334155;'>Cloud Target</th>
                <th style='padding:8px;border:1px solid #334155;'>Upgrade Path</th>
                <th style='padding:8px;border:1px solid #334155;'>Replacement Alternative</th>
                </tr></thead><tbody>
                <tr style='background:#F8FAFC;'><td style='padding:6px 8px;border:1px solid #E2E8F0;'>Windows</td><td style='padding:6px 8px;border:1px solid #E2E8F0;'>AD, .NET, IIS</td><td style='padding:6px 8px;border:1px solid #E2E8F0;color:#1D4ED8;'>Azure</td><td style='padding:6px 8px;border:1px solid #E2E8F0;'>Win 2025 LTSC</td><td style='padding:6px 8px;border:1px solid #E2E8F0;'>Azure PaaS</td></tr>
                <tr style='background:#FFF;'><td style='padding:6px 8px;border:1px solid #E2E8F0;'>RHEL/CentOS</td><td style='padding:6px 8px;border:1px solid #E2E8F0;'>Java, Oracle DB</td><td style='padding:6px 8px;border:1px solid #E2E8F0;color:#1D4ED8;'>AWS</td><td style='padding:6px 8px;border:1px solid #E2E8F0;'>RHEL 9</td><td style='padding:6px 8px;border:1px solid #E2E8F0;'>Amazon Linux 2023</td></tr>
                <tr style='background:#F8FAFC;'><td style='padding:6px 8px;border:1px solid #E2E8F0;'>Ubuntu/Debian</td><td style='padding:6px 8px;border:1px solid #E2E8F0;'>Containers, K8s</td><td style='padding:6px 8px;border:1px solid #E2E8F0;color:#1D4ED8;'>GCP</td><td style='padding:6px 8px;border:1px solid #E2E8F0;'>Ubuntu 24.04 LTS</td><td style='padding:6px 8px;border:1px solid #E2E8F0;'>GKE Autopilot</td></tr>
                <tr style='background:#FFF;'><td style='padding:6px 8px;border:1px solid #E2E8F0;'>Oracle Linux</td><td style='padding:6px 8px;border:1px solid #E2E8F0;'>Oracle DB, PL/SQL</td><td style='padding:6px 8px;border:1px solid #E2E8F0;color:#1D4ED8;'>OCI</td><td style='padding:6px 8px;border:1px solid #E2E8F0;'>OL 9</td><td style='padding:6px 8px;border:1px solid #E2E8F0;'>OCI Autonomous DB</td></tr>
                <tr style='background:#F8FAFC;'><td style='padding:6px 8px;border:1px solid #E2E8F0;'>SLES</td><td style='padding:6px 8px;border:1px solid #E2E8F0;'>SAP, HPC</td><td style='padding:6px 8px;border:1px solid #E2E8F0;color:#1D4ED8;'>Azure/AWS</td><td style='padding:6px 8px;border:1px solid #E2E8F0;'>SLES 15 SP6</td><td style='padding:6px 8px;border:1px solid #E2E8F0;'>Managed SAP</td></tr>
                </tbody></table>
                """, unsafe_allow_html=True)

                st.markdown("**Select your preferred cloud target:**")
                for name, emoji, desc, key in all_profiles:
                    col1, col2 = st.columns([1, 10])
                    with col1:
                        st.checkbox(f"{emoji}", key=f"survey_cloud_{key}", value=False)
                    with col2:
                        st.markdown(f"**{name}**  \n<small style='color:#4B5563;'>{desc}</small>",
                                    unsafe_allow_html=True)

            # ══════════════════════════════════════════════════════════════════
            # CATEGORY 2a: OS & DATABASE UPGRADE MANDATES
            # ══════════════════════════════════════════════════════════════════
            with _t_upgrade:
                st.markdown("""
                <div style="background:#F0FDF4;border:1px solid #A7F3D0;border-radius:8px;
                            padding:0.8rem 1rem;margin-bottom:0.8rem;">
                  <h4 style="margin:0;color:#065F46;">Category 2: OS & Database Upgrade Mandates</h4>
                  <small style="color:#64748B;">Confirm which OS families and databases are in your
                  active landscape. Unchecked items will be excluded from upgrade analysis.</small>
                </div>""", unsafe_allow_html=True)

                st.markdown("##### 🖥️ Operating System Families")
                st.caption("From Agent 1 baseline — tick families present in your environment")
                for fam_name, desc, emoji in family_display:
                    count = len(os_families.get(fam_name, []))
                    if fam_name == "Other":
                        versions_preview = ""
                        default = False
                    else:
                        if count == 0:
                            continue
                        versions_preview = ", ".join(os_families[fam_name][:4])
                        if count > 4:
                            versions_preview += f" + {count - 4} more"
                        default = True
                    col1, col2 = st.columns([1, 4])
                    with col1:
                        st.checkbox(f"{emoji} {fam_name}", value=default, key=f"survey_os_{fam_name}")
                    with col2:
                        st.markdown(
                            f"<small style='color:#6B7280;'>{desc}"
                            + (f" — <em>{versions_preview}</em>" if versions_preview else "")
                            + f" ({count} tracked)</small>", unsafe_allow_html=True)

                st.divider()
                st.markdown("##### 🗄️ Database Families")
                st.caption("Confirm database families in your environment")
                for fam_name, desc, emoji in db_family_display:
                    count = len(db_families.get(fam_name, []))
                    if fam_name == "Other":
                        default = False
                    else:
                        if count == 0:
                            continue
                        default = True
                    versions_preview = ", ".join(db_families.get(fam_name, [])[:3])
                    if count > 3:
                        versions_preview += f" +{count - 3} more"
                    col1, col2 = st.columns([1, 4])
                    with col1:
                        st.checkbox(f"{emoji} {fam_name}", value=default, key=f"survey_db_{fam_name}")
                    with col2:
                        st.markdown(
                            f"<small style='color:#6B7280;'>{desc}"
                            + (f" — <em>{versions_preview}</em>" if versions_preview else "")
                            + f" ({count} tracked)</small>", unsafe_allow_html=True)

            # ══════════════════════════════════════════════════════════════════
            # CATEGORY 2b: REPLACEMENT MANDATES (Web/App/Framework)
            # ══════════════════════════════════════════════════════════════════
            with _t_replace:
                st.markdown("""
                <div style="background:#FEF2F2;border:1px solid #FECACA;border-radius:8px;
                            padding:0.8rem 1rem;margin-bottom:0.8rem;">
                  <h4 style="margin:0;color:#991B1B;">Category 2: Replacement Mandates</h4>
                  <small style="color:#64748B;">Confirm Web Servers, App Servers, and Frameworks
                  in your landscape. EOL items will be flagged for mandatory replacement.</small>
                </div>""", unsafe_allow_html=True)

                st.markdown("##### 🌐 Web Servers")
                wc1, wc2 = st.columns(2)
                for i, ws in enumerate(ws_names):
                    with (wc1 if i % 2 == 0 else wc2):
                        st.checkbox(ws, value=True, key=f"survey_ws_{ws}")

                st.divider()
                st.markdown("##### ⚙️ Application Servers")
                ac1, ac2 = st.columns(2)
                for i, a_s in enumerate(as_names):
                    with (ac1 if i % 2 == 0 else ac2):
                        st.checkbox(a_s, value=True, key=f"survey_as_{a_s}")

                st.divider()
                st.markdown("##### 📦 Frameworks & Runtimes")
                fc1, fc2 = st.columns(2)
                for i, fw in enumerate(fw_names):
                    with (fc1 if i % 2 == 0 else fc2):
                        st.checkbox(fw, value=True, key=f"survey_fw_{fw}")

            # ══════════════════════════════════════════════════════════════════
            # CATEGORY 3: PRINCIPLES & COSTS (preview)
            # ══════════════════════════════════════════════════════════════════
            with _t_policy:
                st.markdown("""
                <div style="background:#FFF7ED;border:1px solid #FED7AA;border-radius:8px;
                            padding:0.8rem 1rem;margin-bottom:0.8rem;">
                  <h4 style="margin:0;color:#92400E;">Category 3: Interactive Principles & Costs</h4>
                  <small style="color:#64748B;">After confirming Categories 1-2, Agent 5 will generate
                  Guiding Principles and cost estimates. Then a Policy Chat will collect your
                  institutional context (compliance, budgets, risk tolerance).</small>
                </div>""", unsafe_allow_html=True)

                st.info(
                    "**What happens next after you confirm:**\n\n"
                    "1. **Guiding Principles Table** — AI-generated upgrade & replacement rules per technology\n"
                    "2. **Cost Estimator** — Upgrade vs Replace vs Do Nothing costs\n"
                    "3. **Migration Wave Planner** — Technologies grouped by urgency\n"
                    "4. **Policy Chat** — Agent 5 asks 10-15 questions about your compliance, budgets, capacity\n"
                    "5. **Final Recommendations** — Cross-referenced with Agent 2's technical analysis\n\n"
                    "*All outputs are exported to Excel, PowerPoint, and PDF.*"
                )

            # ── Confirm All ──────────────────────────────────────────────────
            st.divider()
            if st.button("✅ Confirm All Categories → Generate Guiding Principles",
                         type="primary", use_container_width=True, key="survey_confirm"):
                # Collect OS selections
                os_selected = [fam for fam, _, _ in family_display
                               if st.session_state.get(f"survey_os_{fam}", False) and fam != "Other"]
                st.session_state.a5_landscape_selected = os_selected
                st.session_state.a5_context["os_landscape"] = ", ".join(os_selected)

                # Collect DB selections
                db_selected = [fam for fam, _, _ in db_family_display
                               if st.session_state.get(f"survey_db_{fam}", False) and fam != "Other"]
                st.session_state.a5_db_landscape_selected = db_selected
                st.session_state.a5_context["db_landscape"] = ", ".join(db_selected)

                # Collect Cloud selection
                cloud_sel = None
                for name, emoji, desc, key in all_profiles:
                    if st.session_state.get(f"survey_cloud_{key}", False):
                        cloud_sel = (name, key)
                        break
                if cloud_sel:
                    st.session_state.a5_context["cloud_provider"] = cloud_sel[0]
                    st.session_state.a5_context["cloud_key"] = cloud_sel[1]
                else:
                    st.session_state.a5_context["cloud_provider"] = "No strong preference"
                    st.session_state.a5_context["cloud_key"] = ""

                # Store WS/AS/FW context
                ws_sel = [ws for ws in ws_names if st.session_state.get(f"survey_ws_{ws}", False)]
                as_sel = [a_s for a_s in as_names if st.session_state.get(f"survey_as_{a_s}", False)]
                fw_sel = [fw for fw in fw_names if st.session_state.get(f"survey_fw_{fw}", False)]
                st.session_state.a5_context["ws_landscape"] = ", ".join(ws_sel)
                st.session_state.a5_context["as_landscape"] = ", ".join(as_sel)
                st.session_state.a5_context["fw_landscape"] = ", ".join(fw_sel)

                # Proceed to Guiding Principles generation
                st.session_state.a5_status = "principles_table"
                st.rerun()

        # ── PHASE 0d: GUIDING PRINCIPLES TABLE ──────────────────────────────
        elif a5s == "principles_table":
            from agents.agent_analysis import generate_principles_table

            st.markdown("""
            <div style="background:linear-gradient(135deg,#F0FDF4,#ECFDF5);
                        border:1px solid #A7F3D0;border-radius:12px;
                        padding:1.2rem 1.4rem;margin-bottom:1rem;">
              <h3 style="margin:0 0 0.5rem;color:#065F46;font-size:1.05rem;">
                📋 OS Migration Guiding Principles
              </h3>
              <p style="margin:0;color:#374151;font-size:0.88rem;">
                Based on your <strong>OS landscape</strong> and <strong>cloud profile</strong>,
                here are the guiding principles for your modernization strategy.
              </p>
            </div>""", unsafe_allow_html=True)

            selected_fams = st.session_state.get("a5_landscape_selected", [])
            cloud_name = st.session_state.get("a5_context", {}).get("cloud_provider", "No preference")
            cloud_key = st.session_state.get("a5_context", {}).get("cloud_key", "")

            # Generate or use cached table
            if "a5_principles_table_data" not in st.session_state:
                with st.spinner("🧠 Agent 5 generating deep-dive guiding principles across all categories..."):
                    try:
                        agent5 = PolicyAnalysisAgent(api_key=api_key)
                        table_data = generate_principles_table(
                            selected_fams, cloud_name, cloud_key, agent=agent5,
                            db_df=st.session_state.db_df,
                            ws_df=st.session_state.ws_df,
                            as_df=st.session_state.as_df,
                            fw_df=st.session_state.fw_df)
                    except Exception:
                        table_data = generate_principles_table(
                            selected_fams, cloud_name, cloud_key, agent=None,
                            db_df=st.session_state.db_df,
                            ws_df=st.session_state.ws_df,
                            as_df=st.session_state.as_df,
                            fw_df=st.session_state.fw_df)
                    st.session_state.a5_principles_table_data = table_data
                st.rerun()

            table_data = st.session_state.a5_principles_table_data

            if table_data:
                cat_colors = {
                    "OS": ("#1E3A8A", "🖥️"), "Database": ("#7C2D12", "🗄️"),
                    "Web Server": ("#065F46", "🌐"), "App Server": ("#4338CA", "⚙️"),
                    "Framework": ("#831843", "📦"),
                }
                cat_order = ["OS", "Database", "Web Server", "App Server", "Framework"]

                # Single unified table with category column
                table_html = (
                    "<table style='width:100%;border-collapse:collapse;font-size:0.82rem;table-layout:fixed;'>"
                    "<colgroup>"
                    "<col style='width:10%;'/><col style='width:13%;'/><col style='width:10%;'/>"
                    "<col style='width:33%;'/><col style='width:34%;'/>"
                    "</colgroup>"
                    "<thead><tr style='background:#1E293B;color:white;'>"
                    "<th style='padding:10px;border:1px solid #334155;'>Category</th>"
                    "<th style='padding:10px;border:1px solid #334155;'>Technology</th>"
                    "<th style='padding:10px;border:1px solid #334155;'>Cloud Target</th>"
                    "<th style='padding:10px;border:1px solid #334155;'>Upgrade Principle</th>"
                    "<th style='padding:10px;border:1px solid #334155;'>Replacement Principle</th>"
                    "</tr></thead><tbody>"
                )
                row_idx = 0
                for cat in cat_order:
                    cat_rows = [r for r in table_data if r.get("category", "OS") == cat]
                    if not cat_rows:
                        continue
                    color, emoji = cat_colors.get(cat, ("#374151", "📋"))

                    for i, row in enumerate(cat_rows):
                        bg = "#F8FAFC" if row_idx % 2 == 0 else "#FFFFFF"
                        upg = row.get("upgrade_principle", "")
                        repl = row.get("replacement_principle", "")
                        upg_color = "#166534" if "COTS" in upg else ("#92400E" if "None" in upg else "#1E40AF")
                        repl_color = "#7C2D12" if "Mandatory" in repl or "Migrate" in repl else "#4338CA"
                        tech = row.get("technology", row.get("os_family", ""))

                        # Show category badge on first row of each category
                        if i == 0:
                            cat_cell = (f"<span style='background:{color};color:white;padding:3px 8px;"
                                        f"border-radius:10px;font-size:0.72rem;font-weight:600;'>"
                                        f"{emoji} {cat}</span>")
                        else:
                            cat_cell = ""

                        table_html += (
                            f"<tr style='background:{bg};'>"
                            f"<td style='padding:6px 8px;border:1px solid #E2E8F0;vertical-align:top;'>{cat_cell}</td>"
                            f"<td style='padding:6px 8px;border:1px solid #E2E8F0;font-weight:600;'>{tech}</td>"
                            f"<td style='padding:6px 8px;border:1px solid #E2E8F0;color:#1D4ED8;'>{row.get('cloud_target','')}</td>"
                            f"<td style='padding:6px 8px;border:1px solid #E2E8F0;color:{upg_color};'>{upg}</td>"
                            f"<td style='padding:6px 8px;border:1px solid #E2E8F0;color:{repl_color};'>{repl}</td>"
                            f"</tr>"
                        )
                        row_idx += 1

                table_html += "</tbody></table>"
                st.markdown(table_html, unsafe_allow_html=True)

                # Store in context for the policy chat
                st.session_state.a5_context["principles_table"] = json.dumps(table_data)

                st.divider()

                # Summary metrics
                mc1, mc2, mc3, mc4, mc5 = st.columns(5)
                cats_in_scope = list(set(r.get("category", "OS") for r in table_data))
                with mc1:
                    st.metric("Categories", len(cats_in_scope))
                with mc2:
                    st.metric("Total Technologies", len(table_data))
                with mc3:
                    os_count = sum(1 for r in table_data if r.get("category") == "OS")
                    st.metric("OS Families", os_count)
                with mc4:
                    upgrade_count = sum(1 for r in table_data if "Upgrade" in r.get("upgrade_principle", "") or "COTS" in r.get("upgrade_principle", ""))
                    st.metric("Upgrade Paths", upgrade_count)
                with mc5:
                    replace_count = sum(1 for r in table_data if "Replace" in r.get("replacement_principle", "") or "Migrate" in r.get("replacement_principle", ""))
                    st.metric("Modernization Candidates", replace_count)

                # ══════════════════════════════════════════════════════════════
                # COST ESTIMATOR
                # ══════════════════════════════════════════════════════════════
                from agents.agent_analysis import get_cost_estimates

                st.divider()
                st.markdown("""<div style='background:#FFF7ED;border:1px solid #FED7AA;border-radius:8px;
                    padding:0.8rem 1rem;margin-bottom:0.5rem;'>
                    <h4 style='margin:0;color:#92400E;'>💰 Cost Estimator — Upgrade vs Replace vs Do Nothing</h4>
                </div>""", unsafe_allow_html=True)

                # ── Disclaimer ───────────────────────────────────────────────
                st.markdown("""<div style='background:#FFFBEB;border:1px solid #FDE68A;border-radius:6px;
                    padding:0.5rem 0.8rem;margin-bottom:0.8rem;font-size:0.78rem;color:#92400E;'>
                    ⚠️ <strong>Disclaimer:</strong> All cost figures are <strong>approximate industry estimates</strong>
                    based on publicly available vendor pricing and AI research. Actual costs vary significantly by:
                    enterprise agreement discounts, volume licensing, geographic region, application complexity,
                    and migration team rates. <strong>These estimates are for directional planning only — always
                    validate with vendor quotes and your procurement team before budgeting.</strong>
                    Costs marked <span style='background:#DCFCE7;padding:1px 4px;border-radius:3px;'>AI-enhanced</span>
                    were researched by Agent 5 using current training knowledge.
                </div>""", unsafe_allow_html=True)

                # ── Generate / cache cost data ───────────────────────────────
                if "a5_costed_data" not in st.session_state:
                    with st.spinner("🧠 Agent 5 researching current vendor pricing..."):
                        try:
                            agent5_cost = PolicyAnalysisAgent(api_key=api_key)
                            costed_data = get_cost_estimates([dict(r) for r in table_data], agent=agent5_cost)
                        except Exception:
                            costed_data = get_cost_estimates([dict(r) for r in table_data], agent=None)
                    st.session_state.a5_costed_data = costed_data
                    st.rerun()

                costed_data = st.session_state.a5_costed_data

                # ── Mode toggle: View vs Edit ────────────────────────────────
                cost_mode_col1, cost_mode_col2, cost_mode_col3 = st.columns([2, 2, 6])
                with cost_mode_col1:
                    edit_mode = st.toggle("✏️ Edit Mode", value=False, key="cost_edit_mode")
                with cost_mode_col2:
                    if st.button("🔄 Re-research with AI", key="cost_ai_refresh"):
                        st.session_state.pop("a5_costed_data", None)
                        st.rerun()

                ai_count = sum(1 for r in costed_data if r.get("cost_source") == "AI-enhanced")
                baseline_count = len(costed_data) - ai_count
                with cost_mode_col3:
                    st.caption(f"📊 {ai_count} AI-enhanced · {baseline_count} baseline estimates · {len(costed_data)} total")

                if edit_mode:
                    # ── Editable mode using st.data_editor ───────────────────
                    import pandas as pd
                    cost_df = pd.DataFrame([{
                        "Technology": r.get("technology", r.get("os_family", "")),
                        "Category": r.get("category", ""),
                        "Upgrade Cost": r.get("cost_upgrade", ""),
                        "Replace Cost": r.get("cost_replace", ""),
                        "Do Nothing (Annual)": r.get("cost_do_nothing", ""),
                        "Unit": r.get("cost_unit", ""),
                        "Source": r.get("cost_source", "baseline"),
                    } for r in costed_data])

                    edited_df = st.data_editor(
                        cost_df,
                        use_container_width=True,
                        hide_index=True,
                        disabled=["Technology", "Category", "Source"],
                        column_config={
                            "Technology": st.column_config.TextColumn(width=150),
                            "Category": st.column_config.TextColumn(width=100),
                            "Upgrade Cost": st.column_config.TextColumn("💚 Upgrade Cost", width=180),
                            "Replace Cost": st.column_config.TextColumn("🔵 Replace Cost", width=180),
                            "Do Nothing (Annual)": st.column_config.TextColumn("🔴 Do Nothing", width=180),
                            "Unit": st.column_config.TextColumn(width=90),
                            "Source": st.column_config.TextColumn(width=100),
                        },
                        key="cost_editor"
                    )

                    if st.button("💾 Save Edited Costs", type="primary"):
                        # Merge edits back into costed_data
                        for i, row in edited_df.iterrows():
                            if i < len(costed_data):
                                costed_data[i]["cost_upgrade"] = row["Upgrade Cost"]
                                costed_data[i]["cost_replace"] = row["Replace Cost"]
                                costed_data[i]["cost_do_nothing"] = row["Do Nothing (Annual)"]
                                costed_data[i]["cost_unit"] = row["Unit"]
                                costed_data[i]["cost_source"] = "user-edited"
                        st.session_state.a5_costed_data = costed_data
                        st.success("✅ Costs saved!")
                        st.rerun()
                else:
                    # ── Read-only styled table (single HTML block) ──────────
                    cost_html = (
                        "<table style='width:100%;border-collapse:collapse;font-size:0.8rem;table-layout:fixed;'>"
                        "<colgroup>"
                        "<col style='width:13%;'/><col style='width:8%;'/>"
                        "<col style='width:22%;'/><col style='width:22%;'/><col style='width:22%;'/>"
                        "<col style='width:6%;'/><col style='width:7%;'/>"
                        "</colgroup>"
                        "<thead><tr style='background:#92400E;color:white;'>"
                        "<th style='padding:8px;border:1px solid #FED7AA;'>Technology</th>"
                        "<th style='padding:8px;border:1px solid #FED7AA;'>Category</th>"
                        "<th style='padding:8px;border:1px solid #FED7AA;'>💚 Upgrade Cost</th>"
                        "<th style='padding:8px;border:1px solid #FED7AA;'>🔵 Replace Cost</th>"
                        "<th style='padding:8px;border:1px solid #FED7AA;'>🔴 Do Nothing (Annual)</th>"
                        "<th style='padding:8px;border:1px solid #FED7AA;'>Unit</th>"
                        "<th style='padding:8px;border:1px solid #FED7AA;'>Source</th>"
                        "</tr></thead><tbody>"
                    )
                    for i, row in enumerate(costed_data):
                        bg = "#FFFBEB" if i % 2 == 0 else "#FFF7ED"
                        tech = row.get("technology", row.get("os_family", ""))
                        src = row.get("cost_source", "baseline")
                        src_badge = {"AI-enhanced": "background:#DCFCE7;color:#166534",
                                     "user-edited": "background:#DBEAFE;color:#1E40AF",
                                     "baseline": "background:#F1F5F9;color:#64748B"}.get(src, "")
                        note = row.get("cost_note", "")
                        note_html = f"<br><small style='color:#9CA3AF;'>{note}</small>" if note else ""
                        cost_html += (
                            f"<tr style='background:{bg};'>"
                            f"<td style='padding:6px 8px;border:1px solid #FDE68A;font-weight:600;'>{tech}{note_html}</td>"
                            f"<td style='padding:6px 8px;border:1px solid #FDE68A;'>{row.get('category','')}</td>"
                            f"<td style='padding:6px 8px;border:1px solid #FDE68A;color:#166534;'>{row.get('cost_upgrade','')}</td>"
                            f"<td style='padding:6px 8px;border:1px solid #FDE68A;color:#1D4ED8;'>{row.get('cost_replace','')}</td>"
                            f"<td style='padding:6px 8px;border:1px solid #FDE68A;color:#DC2626;'>{row.get('cost_do_nothing','')}</td>"
                            f"<td style='padding:6px 8px;border:1px solid #FDE68A;'>{row.get('cost_unit','')}</td>"
                            f"<td style='padding:6px 8px;border:1px solid #FDE68A;'>"
                            f"<span style='{src_badge};padding:2px 6px;border-radius:10px;font-size:0.7rem;'>{src}</span></td>"
                            f"</tr>"
                        )
                    cost_html += "</tbody></table>"
                    st.markdown(cost_html, unsafe_allow_html=True)

                # ══════════════════════════════════════════════════════════════
                # MIGRATION WAVE PLANNER
                # ══════════════════════════════════════════════════════════════
                from agents.agent_analysis import assign_migration_waves

                st.divider()
                st.markdown("""<div style='background:#EFF6FF;border:1px solid #93C5FD;border-radius:8px;
                    padding:0.8rem 1rem;margin-bottom:0.8rem;'>
                    <h4 style='margin:0;color:#1E3A8A;'>🌊 Migration Wave Planner</h4>
                    <small style='color:#64748B;'>Technologies grouped by urgency into migration waves with timeline estimates.</small>
                </div>""", unsafe_allow_html=True)

                wave_data = assign_migration_waves(
                    table_data,
                    os_df=st.session_state.os_df, db_df=st.session_state.db_df,
                    ws_df=st.session_state.ws_df, as_df=st.session_state.as_df,
                    fw_df=st.session_state.fw_df)

                wave_colors_ui = {
                    1: ("#DC2626", "#FEE2E2", "🔴"), 2: ("#EA580C", "#FFEDD5", "🟠"),
                    3: ("#CA8A04", "#FEF9C3", "🟡"), 4: ("#16A34A", "#DCFCE7", "🟢"),
                }
                for w in [1, 2, 3, 4]:
                    items = [r for r in wave_data if r.get("wave") == w]
                    if not items:
                        continue
                    color, bg, icon = wave_colors_ui[w]
                    wave_name = items[0].get("wave_name", f"Wave {w}")
                    timeline = items[0].get("timeline", "")

                    st.markdown(
                        f"<div style='background:{bg};border-left:4px solid {color};border-radius:0 8px 8px 0;"
                        f"padding:0.6rem 1rem;margin-bottom:0.5rem;'>"
                        f"<strong style='color:{color};'>{icon} {wave_name}</strong>"
                        f"<span style='float:right;color:{color};font-weight:600;'>{timeline}</span>"
                        f"<br><span style='font-size:0.82rem;color:#374151;'>"
                        + " · ".join(f"<b>{r.get('technology', r.get('os_family',''))}</b> ({r.get('category','')})" for r in items)
                        + f"</span></div>", unsafe_allow_html=True)

                # Wave summary metrics
                wm1, wm2, wm3, wm4 = st.columns(4)
                with wm1: st.metric("🔴 Wave 1 (Critical)", sum(1 for r in wave_data if r.get("wave")==1))
                with wm2: st.metric("🟠 Wave 2 (High)", sum(1 for r in wave_data if r.get("wave")==2))
                with wm3: st.metric("🟡 Wave 3 (Plan)", sum(1 for r in wave_data if r.get("wave")==3))
                with wm4: st.metric("🟢 Wave 4 (Monitor)", sum(1 for r in wave_data if r.get("wave")==4))

                # ══════════════════════════════════════════════════════════════
                # DEPENDENCY MAPPING
                # ══════════════════════════════════════════════════════════════
                from agents.agent_analysis import generate_dependency_map

                st.divider()
                st.markdown("""<div style='background:#F0FDF4;border:1px solid #A7F3D0;border-radius:8px;
                    padding:0.8rem 1rem;margin-bottom:0.8rem;'>
                    <h4 style='margin:0;color:#065F46;'>🔗 Dependency Mapping</h4>
                    <small style='color:#64748B;'>Technologies that must move together due to runtime dependencies.</small>
                </div>""", unsafe_allow_html=True)

                dep_data = generate_dependency_map(table_data)
                if dep_data:
                    for dep in dep_data:
                        deps_text = " → ".join(
                            f"**{d['technology']}** ({d['category']})" for d in dep.get("depends_on", []))
                        st.markdown(
                            f"<div style='background:#F8FAFC;border:1px solid #E2E8F0;border-radius:8px;"
                            f"padding:0.5rem 0.8rem;margin-bottom:0.4rem;font-size:0.82rem;'>"
                            f"⚙️ <strong>{dep['source']}</strong> ({dep['source_category']}) "
                            f"<span style='color:#6B7280;'>depends on</span> {deps_text}"
                            f"<br><small style='color:#9CA3AF;'>{dep.get('note','')}</small>"
                            f"</div>", unsafe_allow_html=True)
                else:
                    st.info("No cross-layer dependencies detected in the current selection.")

                # ══════════════════════════════════════════════════════════════
                # COMPLIANCE CROSSWALK
                # ══════════════════════════════════════════════════════════════
                from agents.agent_analysis import generate_compliance_crosswalk

                st.divider()
                st.markdown("""<div style='background:#FEF2F2;border:1px solid #FECACA;border-radius:8px;
                    padding:0.8rem 1rem;margin-bottom:0.8rem;'>
                    <h4 style='margin:0;color:#991B1B;'>🛡️ Compliance Crosswalk</h4>
                    <small style='color:#64748B;'>EOL status mapped against PCI DSS, HIPAA, SOX, NIST, ISO 27001.</small>
                </div>""", unsafe_allow_html=True)

                compliance_data = generate_compliance_crosswalk(wave_data)
                violations = [r for r in compliance_data if r.get("compliance_status") == "VIOLATION"]
                at_risk = [r for r in compliance_data if r.get("compliance_status") == "WARNING"]
                compliant_items = [r for r in compliance_data if r.get("compliance_status") == "COMPLIANT"]

                cc1, cc2, cc3 = st.columns(3)
                with cc1: st.metric("❌ Violations", len(violations))
                with cc2: st.metric("⚠️ At Risk", len(at_risk))
                with cc3: st.metric("✅ Compliant", len(compliant_items))

                if violations:
                    for v in violations:
                        frameworks = " · ".join(
                            f"**{cv['framework']}** ({cv['status']})" for cv in v.get("compliance_violations", []))
                        st.markdown(
                            f"<div style='background:#FEE2E2;border-left:4px solid #DC2626;border-radius:0 6px 6px 0;"
                            f"padding:0.4rem 0.8rem;margin-bottom:0.3rem;font-size:0.82rem;'>"
                            f"❌ <strong>{v.get('technology', v.get('os_family',''))}</strong> ({v.get('category','')}) "
                            f"— {frameworks}</div>", unsafe_allow_html=True)

                if at_risk:
                    with st.expander(f"⚠️ {len(at_risk)} technologies at risk", expanded=False):
                        for w in at_risk:
                            frameworks = " · ".join(
                                f"{cw['framework']}" for cw in w.get("compliance_warnings", []))
                            st.markdown(f"- **{w.get('technology','')}** ({w.get('category','')}) — approaching: {frameworks}")

                # ══════════════════════════════════════════════════════════════
                # POWERPOINT EXPORT
                # ══════════════════════════════════════════════════════════════
                from agents.agent_analysis import generate_pptx

                st.divider()

                # ── Excel export with all recommendations ────────────────────
                try:
                    import io
                    from openpyxl import Workbook
                    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

                    wb = Workbook()

                    # Sheet 1: Guiding Principles
                    ws1 = wb.active
                    ws1.title = "Guiding Principles"
                    hdr_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
                    hdr_font = Font(bold=True, color="FFFFFF", size=10)
                    thin_border = Border(
                        left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))

                    gp_headers = ["Category", "Technology", "Cloud Target", "Upgrade Principle", "Replacement Principle"]
                    for j, h in enumerate(gp_headers, 1):
                        cell = ws1.cell(row=1, column=j, value=h)
                        cell.fill = hdr_fill; cell.font = hdr_font; cell.border = thin_border
                        cell.alignment = Alignment(wrap_text=True)
                    for i, row in enumerate(table_data, 2):
                        vals = [row.get("category",""), row.get("technology", row.get("os_family","")),
                                row.get("cloud_target",""), row.get("upgrade_principle",""),
                                row.get("replacement_principle","")]
                        for j, v in enumerate(vals, 1):
                            cell = ws1.cell(row=i, column=j, value=v)
                            cell.border = thin_border; cell.alignment = Alignment(wrap_text=True)
                    ws1.column_dimensions['A'].width = 14; ws1.column_dimensions['B'].width = 20
                    ws1.column_dimensions['C'].width = 16; ws1.column_dimensions['D'].width = 50
                    ws1.column_dimensions['E'].width = 50

                    # Sheet 2: Cost Estimates
                    ws2 = wb.create_sheet("Cost Estimates")
                    cost_fill = PatternFill(start_color="92400E", end_color="92400E", fill_type="solid")
                    cost_headers = ["Technology", "Category", "Upgrade Cost", "Replace Cost", "Do Nothing (Annual)", "Unit", "Source"]
                    for j, h in enumerate(cost_headers, 1):
                        cell = ws2.cell(row=1, column=j, value=h)
                        cell.fill = cost_fill; cell.font = hdr_font; cell.border = thin_border
                    for i, row in enumerate(costed_data, 2):
                        vals = [row.get("technology", row.get("os_family","")), row.get("category",""),
                                row.get("cost_upgrade",""), row.get("cost_replace",""),
                                row.get("cost_do_nothing",""), row.get("cost_unit",""),
                                row.get("cost_source","baseline")]
                        for j, v in enumerate(vals, 1):
                            cell = ws2.cell(row=i, column=j, value=v)
                            cell.border = thin_border; cell.alignment = Alignment(wrap_text=True)
                    for col, w in [('A',20),('B',12),('C',25),('D',25),('E',25),('F',10),('G',12)]:
                        ws2.column_dimensions[col].width = w

                    # Sheet 3: Migration Waves
                    ws3 = wb.create_sheet("Migration Waves")
                    wave_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
                    wave_headers = ["Wave", "Timeline", "Urgency", "Technology", "Category", "Risk Score", "Days to EOL"]
                    for j, h in enumerate(wave_headers, 1):
                        cell = ws3.cell(row=1, column=j, value=h)
                        cell.fill = wave_fill; cell.font = hdr_font; cell.border = thin_border
                    for i, row in enumerate(wave_data, 2):
                        vals = [row.get("wave_name",""), row.get("timeline",""), row.get("urgency",""),
                                row.get("technology", row.get("os_family","")), row.get("category",""),
                                row.get("risk_score",""), row.get("days_eol","")]
                        for j, v in enumerate(vals, 1):
                            cell = ws3.cell(row=i, column=j, value=v)
                            cell.border = thin_border
                    for col, w in [('A',25),('B',15),('C',12),('D',20),('E',14),('F',12),('G',12)]:
                        ws3.column_dimensions[col].width = w

                    # Sheet 4: Compliance Crosswalk
                    ws4 = wb.create_sheet("Compliance Crosswalk")
                    comp_fill = PatternFill(start_color="991B1B", end_color="991B1B", fill_type="solid")
                    comp_headers = ["Technology", "Category", "Compliance Status", "Violations", "At Risk"]
                    for j, h in enumerate(comp_headers, 1):
                        cell = ws4.cell(row=1, column=j, value=h)
                        cell.fill = comp_fill; cell.font = hdr_font; cell.border = thin_border
                    for i, row in enumerate(compliance_data, 2):
                        violations_str = "; ".join(f"{v['framework']}: {v['status']}" for v in row.get("compliance_violations",[]))
                        warnings_str = "; ".join(f"{w['framework']}" for w in row.get("compliance_warnings",[]))
                        vals = [row.get("technology", row.get("os_family","")), row.get("category",""),
                                row.get("compliance_status",""), violations_str, warnings_str]
                        for j, v in enumerate(vals, 1):
                            cell = ws4.cell(row=i, column=j, value=v)
                            cell.border = thin_border; cell.alignment = Alignment(wrap_text=True)
                    for col, w in [('A',20),('B',14),('C',16),('D',50),('E',40)]:
                        ws4.column_dimensions[col].width = w

                    # Sheet 5: Dependencies
                    ws5 = wb.create_sheet("Dependencies")
                    dep_fill = PatternFill(start_color="065F46", end_color="065F46", fill_type="solid")
                    dep_headers = ["Source Technology", "Source Category", "Depends On", "Dependency Type", "Note"]
                    for j, h in enumerate(dep_headers, 1):
                        cell = ws5.cell(row=1, column=j, value=h)
                        cell.fill = dep_fill; cell.font = hdr_font; cell.border = thin_border
                    row_num = 2
                    for dep in dep_data:
                        for d in dep.get("depends_on", []):
                            vals = [dep.get("source",""), dep.get("source_category",""),
                                    d.get("technology",""), d.get("type",""), dep.get("note","")]
                            for j, v in enumerate(vals, 1):
                                cell = ws5.cell(row=row_num, column=j, value=v)
                                cell.border = thin_border; cell.alignment = Alignment(wrap_text=True)
                            row_num += 1
                    for col, w in [('A',20),('B',16),('C',20),('D',14),('E',50)]:
                        ws5.column_dimensions[col].width = w

                    excel_buf = io.BytesIO()
                    wb.save(excel_buf)
                    excel_buf.seek(0)
                    excel_bytes = excel_buf.getvalue()
                    excel_fname = f"Migration_Recommendations_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
                except Exception:
                    excel_bytes = None

                # ── Download buttons ─────────────────────────────────────────
                dl_a, dl_b = st.columns(2)
                with dl_a:
                    if excel_bytes:
                        st.download_button(
                            "📥 Download Excel (All Recommendations)",
                            data=excel_bytes, file_name=excel_fname,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True, type="primary")
                with dl_b:
                    try:
                        pptx_bytes = generate_pptx(
                            table_data, wave_data, compliance_data, dep_data,
                            cloud_name, selected_fams)
                        pptx_fname = f"Migration_Strategy_{datetime.now().strftime('%Y%m%d_%H%M')}.pptx"
                        st.download_button(
                            "📊 Download PowerPoint Deck",
                            data=pptx_bytes, file_name=pptx_fname,
                            mime="application/vnd.openxmlformats-officedocument.presentationml.presentation",
                            use_container_width=True)
                    except Exception:
                        pass

                # ── FYI note + auto-proceed ──────────────────────────────────
                st.markdown("""<div style='background:#F0FDF4;border:1px solid #BBF7D0;border-radius:8px;
                    padding:0.6rem 1rem;margin-top:0.5rem;font-size:0.85rem;color:#166534;'>
                    ℹ️ <strong>For Your Information:</strong> The above principles, costs, waves, dependencies,
                    and compliance mapping are based on your OS/DB landscape and cloud selection.
                    Proceeding to Policy Chat where Agent 5 will use this context for tailored recommendations.
                </div>""", unsafe_allow_html=True)

                dl1, dl2 = st.columns([3, 1])
                with dl1:
                    if st.button("➡️ Proceed to Policy Chat", type="primary", use_container_width=True):
                        st.session_state.a5_status = "chatting"
                        st.rerun()
                with dl2:
                    if st.button("🔄 Regenerate", use_container_width=True):
                        st.session_state.pop("a5_principles_table_data", None)
                        st.session_state.pop("a5_costed_data", None)
                        st.rerun()
            else:
                st.warning("No OS families selected. Go back to the Landscape Survey.")
                if st.button("⬅️ Back to Landscape Survey"):
                    st.session_state.a5_status = "landscape"
                    st.rerun()

        # ── PHASE 1: CHAT ─────────────────────────────────────────────────────
        elif a5s in ("chatting",):

            # Import SQLite helpers from agent
            from agents.agent_analysis import (
                _save_message, _load_messages, _save_session_context,
                _list_sessions, _delete_session
            )

            # Show past sessions in expander
            sessions = _list_sessions()
            if sessions:
                with st.expander(f"📂 Past conversations ({len(sessions)})", expanded=False):
                    for sess in sessions:
                        sc1, sc2, sc3 = st.columns([3, 1, 1])
                        with sc1:
                            st.markdown(
                                f"**{sess['session']}** — {sess['summary'][:60]}  \n"
                                f"<small>{sess['updated']} · {sess['status']}</small>",
                                unsafe_allow_html=True
                            )
                        with sc2:
                            if st.button("▶ Resume", key=f"res_{sess['session']}"):
                                st.session_state.a5_session_id = sess["session"]
                                st.session_state.a5_status     = sess["status"] if sess["status"] in ("chatting","ready","done") else "chatting"
                                st.rerun()
                        with sc3:
                            if st.button("🗑 Delete", key=f"del_{sess['session']}"):
                                _delete_session(sess["session"])
                                st.rerun()

            # Create session on first entry to chat phase
            session_id = PolicyAnalysisAgent.get_or_create_session()
            if not st.session_state.get("a5_opening_pending") and not _load_messages(session_id):
                st.session_state.a5_opening_pending = True
            messages   = _load_messages(session_id)

            # ── Show landscape context if available ──────────────────────────
            ls_selected = st.session_state.get("a5_landscape_selected", [])
            if ls_selected:
                active_fams = [f for f in ls_selected if f != "Other"]
                if active_fams:
                    st.markdown(
                        f"<div style='background:#EFF6FF;border:1px solid #BFDBFE;border-radius:8px;"
                        f"padding:0.5rem 0.8rem;margin-bottom:0.8rem;font-size:0.82rem;color:#1E40AF;'>"
                        f"🗺️ <strong>OS Landscape:</strong> {' · '.join(active_fams)}"
                        f"</div>", unsafe_allow_html=True)

            # ── Welcome / guidance panel (shown only when chat is empty) ──────
            if not messages:
                st.markdown("""
                <div style="background:linear-gradient(135deg,#EFF6FF,#F0FDF4);
                            border:1px solid #BFDBFE;border-radius:12px;
                            padding:1.4rem 1.6rem;margin-bottom:1.2rem;">
                  <h3 style="margin:0 0 0.6rem;color:#1E40AF;font-size:1.1rem;">
                    👋 Welcome to Agent 5 — Your Migration Policy Advisor
                  </h3>
                  <p style="margin:0 0 0.8rem;color:#374151;font-size:0.9rem;">
                    Agent 5 will guide you through a <strong>natural conversation</strong>
                    to define your organisation's migration policy for
                    <strong>Apr 2026 → Jun 2028</strong>. It adapts to your answers and
                    <strong>searches the internet in real-time</strong> for pricing,
                    ESU costs, migration guides and upgrade paths.
                  </p>
                  <div style="display:grid;grid-template-columns:1fr 1fr;gap:0.6rem;margin-bottom:0.8rem;">
                    <div style="background:white;border-radius:8px;padding:0.6rem 0.8rem;border:1px solid #E5E7EB;">
                      <strong style="color:#1D4ED8;font-size:0.82rem;">💬 What to expect</strong><br>
                      <span style="font-size:0.8rem;color:#6B7280;">
                        10–15 focused questions about risk tolerance, budgets,
                        OS/DB migration paths, compliance scope and execution capacity.
                      </span>
                    </div>
                    <div style="background:white;border-radius:8px;padding:0.6rem 0.8rem;border:1px solid #E5E7EB;">
                      <strong style="color:#059669;font-size:0.82rem;">🔍 Real-time searches</strong><br>
                      <span style="font-size:0.8rem;color:#6B7280;">
                        Ask about ESU costs, cloud DB pricing, RHEL subscriptions,
                        Oracle support fees or any migration guide — it searches live.
                      </span>
                    </div>
                    <div style="background:white;border-radius:8px;padding:0.6rem 0.8rem;border:1px solid #E5E7EB;">
                      <strong style="color:#7C3AED;font-size:0.82rem;">⚖️ Guiding Principles</strong><br>
                      <span style="font-size:0.8rem;color:#6B7280;">
                        Your answers become 8–10 named Guiding Principles (GP-01…)
                        that govern every migration recommendation.
                      </span>
                    </div>
                    <div style="background:white;border-radius:8px;padding:0.6rem 0.8rem;border:1px solid #E5E7EB;">
                      <strong style="color:#DC2626;font-size:0.82rem;">📊 Final Recommendations</strong><br>
                      <span style="font-size:0.8rem;color:#6B7280;">
                        Cross-references Agent 2's technical advice with your policy
                        to produce a Final Recommendations sheet in Excel.
                      </span>
                    </div>
                  </div>
                  <p style="margin:0;color:#6B7280;font-size:0.8rem;">
                    💡 <strong>Tip:</strong> Answer in your own words — you don't need to follow a format.
                    You can also ask Agent 5 questions at any time, e.g.
                    <em>"What does Windows Server 2016 ESU cost?"</em> or
                    <em>"What's the best path to migrate Oracle to PostgreSQL?"</em>
                  </p>
                </div>
                """, unsafe_allow_html=True)

                # ── Starter prompts ───────────────────────────────────────────
                st.markdown("**🚀 Start with one of these, or type your own:**")
                starter_cols = st.columns(2)
                starters = [
                    ("🏢 Enterprise context",
                     "We are a large enterprise running a mix of Windows Server, RHEL, Oracle and SQL Server. We have PCI DSS compliance requirements and our priority is zero EOL risk for Tier-1 systems."),
                    ("💰 Cost-focused",
                     "Our budget is constrained. We need to know the real cost of extending support vs migrating. Can you help me understand the ESU pricing for Windows Server 2016 and SQL Server 2017?"),
                    ("☁️ Cloud migration",
                     "We are planning to migrate our SQL Server and Oracle databases to the cloud. Our preferred provider is Azure. What should I consider for the migration policy?"),
                    ("🔒 Compliance-led",
                     "We are under HIPAA and SOX compliance. Our risk tolerance for EOL software is zero for regulated systems. We have a small migration team of 5 people."),
                ]
                for i, (label, prompt_text) in enumerate(starters):
                    with starter_cols[i % 2]:
                        if st.button(label, key=f"starter_{i}", width="stretch",
                                     help=prompt_text[:80] + "..."):
                            _save_message(session_id, "user", prompt_text)
                            with st.spinner("🧠 Agent 5 is preparing your personalised policy session..."):
                                agent5   = PolicyAnalysisAgent(api_key=api_key)
                                all_msgs = _load_messages(session_id)
                                reply    = agent5.chat(all_msgs)
                                _save_message(session_id, "assistant", reply)
                                _save_session_context(session_id, {}, "", "chatting")
                            st.rerun()

                st.divider()

                # ── Generate opening message if not yet done ──────────────────
                if st.session_state.get("a5_opening_pending"):
                    with st.spinner("🧠 Agent 5 is preparing your session..."):
                        agent5  = PolicyAnalysisAgent(api_key=api_key)
                        opening = agent5.chat([])
                        _save_message(session_id, "assistant", opening)
                        st.session_state.a5_opening_pending = False
                    st.rerun()
                else:
                    if st.button("▶ Let Agent 5 start the conversation", type="primary",
                                 width="stretch"):
                        st.session_state.a5_opening_pending = True
                        st.rerun()

            else:
                # ── Show session info bar ─────────────────────────────────────
                st.markdown(
                    f"<div style='background:#F0FDF4;border:1px solid #BBF7D0;border-radius:8px;"
                    f"padding:0.5rem 1rem;margin-bottom:0.8rem;font-size:0.82rem;color:#166534;'>"
                    f"💬 Session <strong>{session_id}</strong> · "
                    f"{len(messages)} message{'s' if len(messages) != 1 else ''} · "
                    f"🔍 Agent 5 searches the web in real-time for costs, pricing and migration guides"
                    f"</div>",
                    unsafe_allow_html=True
                )

                # ── Render chat history ───────────────────────────────────────
                for msg in messages:
                    content = msg["content"]
                    # Hide raw JSON responses from display
                    if msg["role"] == "assistant" and content.strip().startswith("{") and '"ready"' in content:
                        continue  # Skip JSON-only messages
                    with st.chat_message(msg["role"],
                                         avatar="🧠" if msg["role"] == "assistant" else "👤"):
                        st.markdown(content)

            # ── User chat input (always shown when chatting) ──────────────────
            if messages or not st.session_state.get("a5_opening_pending"):
                if user_input := st.chat_input(
                    "Ask about ESU costs, migration options, compliance, or answer Agent 5's questions..."
                ):
                    _save_message(session_id, "user", user_input)
                    with st.chat_message("user", avatar="👤"):
                        st.markdown(user_input)

                    with st.chat_message("assistant", avatar="🧠"):
                        with st.spinner("🔍 Agent 5 is thinking..."):
                            agent5   = PolicyAnalysisAgent(api_key=api_key)
                            all_msgs = _load_messages(session_id)
                            reply    = agent5.chat(all_msgs)

                        # Adaptive completion — accepts 4+ exchanges, fills defaults for gaps
                        done, context, summary, inferred = agent5.is_conversation_complete(
                            reply, message_count=len(all_msgs) + 1)

                        if done:
                            inferred_note = ""
                            if inferred:
                                inferred_note = (
                                    f"\n\n💡 **Topics using defaults** (not explicitly discussed): "
                                    f"{', '.join(inferred)}. "
                                    f"These can be refined by continuing the conversation."
                                )
                            complete_msg = (
                                f"✅ **I have enough policy context to proceed.**\n\n"
                                f"**Policy Summary:** {summary}"
                                f"{inferred_note}\n\n"
                                f"Moving to Phase 2 — generating Guiding Principles..."
                            )
                            st.markdown(complete_msg)
                            _save_message(session_id, "assistant", complete_msg)
                            _save_session_context(session_id, context, summary, "principles")
                            st.session_state.a5_context = context
                            st.session_state.a5_status  = "principles"
                        else:
                            # If reply is raw JSON, show a friendly message instead
                            display_reply = reply
                            if reply.strip().startswith("{") and '"ready"' in reply:
                                display_reply = ("I'm gathering your policy context. "
                                    "Let's continue — could you share more about your "
                                    "migration priorities and constraints?")
                            st.markdown(display_reply)
                            _save_message(session_id, "assistant", display_reply)
                            _save_session_context(session_id, {}, "", "chatting")
                    st.rerun()

            # ── Topic coverage + controls ─────────────────────────────────────
            if messages:
                TOPIC_LABELS = {
                    "eol_tolerance": "EOL risk tolerance",  "min_runway": "Support runway",
                    "esu_budget": "ESU budget",             "compliance": "Compliance scope",
                    "windows_path": "Windows path",         "linux_path": "Linux/Unix path",
                    "db_eol_path": "DB EOL path",           "oracle_stance": "Oracle stance",
                    "cloud_provider": "Cloud provider",     "legacy_db": "Legacy DBs",
                    "capacity": "Migration capacity",        "criticality": "System priority",
                    "rollback": "Rollback policy",
                }
                saved_ctx = st.session_state.get("a5_context", {})
                explicit  = [k for k, l in TOPIC_LABELS.items()
                             if saved_ctx.get(k,"").strip() and not saved_ctx.get(k,"").startswith("[Default]")]
                msg_count = len(messages)
                pct       = int(len(explicit) / len(TOPIC_LABELS) * 100)

                if msg_count >= 8:
                    colour = "#F0FDF4"; border = "#BBF7D0"; text_c = "#166534"
                    note   = f"✅ {len(explicit)}/{len(TOPIC_LABELS)} topics explicitly covered"
                else:
                    colour = "#FFF7ED"; border = "#FED7AA"; text_c = "#92400E"
                    note   = f"💬 {len(explicit)}/{len(TOPIC_LABELS)} topics covered · keep chatting or proceed when ready"

                st.markdown(
                    f"<div style='background:{colour};border:1px solid {border};"
                    f"border-radius:8px;padding:0.5rem 0.9rem;font-size:0.8rem;color:{text_c};'>"
                    f"{note} · {msg_count} messages<br>"
                    f"<span style='color:#6B7280;font-size:0.75rem;'>"
                    f"Missing topics will use enterprise defaults — you can refine anytime</span>"
                    f"</div>", unsafe_allow_html=True)

                st.divider()
                col_proceed, col_reset = st.columns([2, 1])
                with col_proceed:
                    # Allow manual proceed after just 4 exchanges (8 messages)
                    if len(messages) >= 8:
                        if st.button("➡️ Proceed to Guiding Principles",
                                     type="primary", width="stretch"):
                            agent5     = PolicyAnalysisAgent(api_key=api_key)
                            all_msgs   = _load_messages(session_id)
                            ctx_prompt = (
                                "Extract all policy context discussed so far as JSON with these keys: "
                                "eol_tolerance, min_runway, esu_budget, compliance, "
                                "windows_path, linux_path, db_eol_path, oracle_stance, "
                                "cloud_provider, legacy_db, capacity, criticality, rollback. "
                                "For any topic NOT discussed, use '[Default] <sensible enterprise default>'. "
                                "Return ONLY the JSON object."
                            )
                            try:
                                r = agent5.client.chat.completions.create(
                                    model="gpt-4o-mini", max_tokens=700,
                                    messages=all_msgs + [{"role":"user","content":ctx_prompt}]
                                )
                                t = r.choices[0].message.content.strip()
                                s, e = t.find("{"), t.rfind("}")
                                context = json.loads(t[s:e+1]) if s != -1 and e > s else {}
                            except Exception:
                                context = {"note": "Extracted from conversation"}
                            _save_session_context(session_id, context, "Manual proceed", "principles")
                            st.session_state.a5_context = context
                            st.session_state.a5_status  = "principles"
                            st.rerun()
                    else:
                        remaining = max(0, 8 - len(messages))
                        st.caption(f"💬 {remaining} more message(s) before you can proceed manually")
                with col_reset:
                    if st.button("🔄 New Conversation", width="stretch"):
                        PolicyAnalysisAgent.reset()
                        st.rerun()

        # ── PHASE 2: Generate Principles ──────────────────────────────────────
        elif a5s == "principles":
            st.subheader("⚖️ Generating Guiding Principles from your policy conversation...")
            with st.spinner("OpenAI is synthesising your answers into Guiding Principles..."):
                session_id = PolicyAnalysisAgent.get_or_create_session()
                agent5     = PolicyAnalysisAgent(api_key=api_key)
                principles = agent5.generate_principles(
                    st.session_state.a5_context,
                    session_id
                )
                st.session_state.a5_principles = principles
                st.session_state.a5_status     = "costing"
            st.rerun()

        # ── PHASE 3: Cost Intelligence ─────────────────────────────────────────
        elif a5s == "costing":
            st.subheader("💰 Fetching Live Vendor Pricing...")
            st.caption("Agent 5 is searching the web for current ESU costs, cloud DB pricing, RHEL subscription rates and migration tool costs.")
            prog = st.progress(0, text="Searching live pricing...")
            stat = st.empty()
            def cost_cb(pct, msg):
                prog.progress(min(pct, 1.0), text=msg)
                stat.info(f"⏳ {msg}")
            agent5 = PolicyAnalysisAgent(api_key=api_key)
            costs  = agent5.fetch_costs(progress_cb=cost_cb)
            st.session_state.a5_costs  = costs
            st.session_state.a5_status = "ready"
            st.rerun()

        # ── PHASE 3 READY: Show principles + confirm ──────────────────────────
        elif a5s == "ready":
            principles = st.session_state.a5_principles
            context    = st.session_state.a5_context

            # Show policy summary from chat
            if context:
                with st.expander("📋 Policy Context from conversation", expanded=False):
                    for k, v in context.items():
                        is_default = str(v).startswith("[Default]")
                        icon = "⚙️" if is_default else "✅"
                        label = k.replace("_"," ").title()
                        st.markdown(f"{icon} **{label}:** {v}"
                                    + (" *(enterprise default — not discussed)*" if is_default else ""))

            with st.expander(f"⚖️ {len(principles)} Guiding Principles generated", expanded=True):
                c1, c2 = st.columns(2)
                cat_colors = {"Risk":"#EF4444","Budget":"#F59E0B","OS":"#3B82F6",
                              "Database":"#8B5CF6","Execution":"#10B981"}
                for i, gp in enumerate(principles):
                    cat   = gp.get("category","")
                    color = cat_colors.get(cat, "#6B7280")
                    with (c1 if i % 2 == 0 else c2):
                        st.markdown(
                            f"<div style='border-left:3px solid {color};padding:8px 12px;"
                            f"border-radius:0 6px 6px 0;background:rgba(0,0,0,0.03);margin-bottom:8px;'>"
                            f"<strong>{gp.get('code','?')}: {gp.get('title','?')}</strong> "
                            f"<span style='font-size:0.7rem;color:{color};background:rgba(0,0,0,0.06);"
                            f"padding:1px 6px;border-radius:8px;'>{cat}</span><br>"
                            f"<span style='font-size:0.82rem;'>{gp.get('rule','')}</span><br>"
                            f"<span style='font-size:0.75rem;color:#6B7280;'>Trigger: {gp.get('trigger','')}</span>"
                            f"</div>", unsafe_allow_html=True)

            with st.expander(f"💰 Vendor Cost Intelligence ({len(st.session_state.a5_costs)} sources)", expanded=False):
                for vendor, summary in st.session_state.a5_costs.items():
                    st.markdown(f"**{vendor}:** {summary}")

            st.divider()
            st.subheader("🚀 Ready to Generate Final Recommendations")

            a2_os = (st.session_state.os_df["Recommendation"] != "").sum()
            a2_db = (st.session_state.db_df["Recommendation"] != "").sum()
            total_rows = len(st.session_state.os_df) + len(st.session_state.db_df)

            if a2_os + a2_db == 0:
                st.warning("⚠️ Run Agent 2 first — Final Recommendations are richer when Agent 2's technical advice is available.")
            else:
                st.success(f"✅ Agent 2 recommendations available: {a2_os} OS + {a2_db} DB rows")

            st.info(
                f"Agent 5 will now cross-reference **{total_rows} entries** against your "
                f"**{len(principles)} Guiding Principles** and Agent 2's recommendations to produce "
                f"a **Final Recommendation** column — this will appear as a separate sheet in the Excel export."
            )

            if st.button("▶ Generate Final Recommendations", type="primary", width="stretch"):
                st.session_state.a5_status = "analysing"
                st.rerun()
            if st.button("💬 Continue the conversation / refine policy", width="stretch"):
                st.session_state.a5_status = "chatting"
                st.rerun()
            if st.button("🔄 Start Over", width="stretch"):
                PolicyAnalysisAgent.reset()
                st.rerun()

        # ── PHASE 4: Analysis ─────────────────────────────────────────────────
        elif a5s == "analysing":
            st.subheader("🧠 Agent 5 — Generating Final Recommendations...")

            if "a5_log" not in st.session_state:
                st.session_state.a5_log = []

            def _log(icon, msg):
                st.session_state.a5_log.append(f"{icon} {msg}")

            log_box = st.container()
            with log_box:
                for entry in st.session_state.a5_log:
                    if entry.startswith("❌"): st.error(entry)
                    elif entry.startswith("⚠️"): st.warning(entry)
                    elif entry.startswith("✅"): st.success(entry)
                    else: st.info(entry)

            prog_bar = st.progress(0, text="Starting...")
            st.divider()

            # Pre-flight
            if not st.session_state.get("a5_preflight_done"):
                _log("🔌", "**Step 1/4 — Testing OpenAI API connection...**")
                try:
                    from openai import OpenAI as _OAI
                    _client = _OAI(api_key=api_key)
                    # Try models in order until one works
                    _model_used = None
                    for _m in ["gpt-4o-mini", "gpt-3.5-turbo", "gpt-3.5-turbo-0125"]:
                        try:
                            _resp = _client.chat.completions.create(
                                model=_m, max_tokens=10,
                                messages=[{"role":"user","content":"Reply: READY"}]
                            )
                            _model_used = _m
                            st.session_state["_openai_model"] = _m
                            break
                        except Exception as _me:
                            if "not found" in str(_me).lower() or "404" in str(_me):
                                continue
                            raise _me
                    if not _model_used:
                        raise Exception("No supported model found on this OpenAI account")
                    _reply = _resp.choices[0].message.content.strip()
                    _log("✅", f"**OpenAI API connected** — model: `{_model_used}` · response: **{_reply}**")
                    st.session_state.a5_preflight_done = True
                except Exception as _ex:
                    _log("❌", f"**OpenAI API FAILED** — `{str(_ex)}`\n\nCheck: platform.openai.com/usage")
                    with log_box: st.error(st.session_state.a5_log[-1])
                    st.stop()
                st.rerun()

            # Analysis steps: OS → DB → WS → AS → FW
            _a5_steps = [
                ("a5_os_done", "os_df", "OS", "analyse_os", 2),
                ("a5_db_done", "db_df", "DB", "analyse_db", 3),
                ("a5_ws_done", "ws_df", "WS", "analyse_ws", 4),
                ("a5_as_done", "as_df", "AS", "analyse_as", 5),
                ("a5_fw_done", "fw_df", "FW", "analyse_fw", 6),
            ]
            total_steps = len(_a5_steps) + 2  # +1 preflight, +1 wrapup = 7

            for done_key, df_key, kind_label, method_name, step_num in _a5_steps:
                if not st.session_state.get(done_key):
                    agent5     = PolicyAnalysisAgent(api_key=api_key)
                    principles = st.session_state.a5_principles
                    costs      = st.session_state.a5_costs
                    context    = st.session_state.a5_context
                    cur_df     = st.session_state[df_key]
                    _log("🧠", f"**Step {step_num}/{total_steps} — Final Recommendations for {len(cur_df)} {kind_label} entries...**")

                    _live = st.empty()
                    frac_start = (step_num - 2) / (total_steps - 2)
                    frac_end   = (step_num - 1) / (total_steps - 2)
                    def _make_cb(fs=frac_start, fe=frac_end):
                        def cb(pct, msg):
                            prog_bar.progress(min(fs + pct * (fe - fs), 1.0), text=msg)
                            _live.info(f"⏳ {msg}")
                        return cb

                    analyse_fn = getattr(agent5, method_name)
                    new_df = analyse_fn(cur_df, principles, costs, context, _make_cb())
                    st.session_state[df_key] = new_df
                    st.session_state[done_key] = True
                    # Persist OS/DB to SQLite
                    if df_key == "os_df":
                        save_os_df(new_df)
                    elif df_key == "db_df":
                        save_db_df(new_df)

                    if "Analysis Source" in new_df.columns:
                        ai_c = (new_df["Analysis Source"] == "OpenAI").sum()
                        rb_c = (new_df["Analysis Source"] == "Rule-based").sum()
                        _log("✅" if rb_c == 0 else "⚠️",
                             f"**{kind_label} done** — OpenAI: {ai_c} rows ✅" +
                             (f" | Rule-based fallback: {rb_c} rows" if rb_c else ""))
                    st.rerun()

            _log("🏁", f"**Step {total_steps}/{total_steps} — Final Recommendations complete for all categories.**")
            st.session_state.a5_status         = "done"
            st.session_state.a5_preflight_done = False
            st.session_state.a5_log            = []
            st.rerun()

        # ── PHASE 5: DONE ─────────────────────────────────────────────────────
        elif a5s == "done":
            st.success(
                "✅ **Agent 5 complete!** "
                "'Final Recommendation' and 'Final Verdict' columns added to OS and DB tabs. "
                "A dedicated **Final Recommendations** sheet is included in the Excel export."
            )

            os_df_now = st.session_state.os_df
            db_df_now = st.session_state.db_df

            # AI vs rule-based breakdown
            if "Analysis Source" in os_df_now.columns or "Analysis Source" in db_df_now.columns:
                ai_os = (os_df_now.get("Analysis Source", pd.Series(dtype=str)) == "OpenAI").sum() if "Analysis Source" in os_df_now.columns else 0
                rb_os = (os_df_now.get("Analysis Source", pd.Series(dtype=str)) == "Rule-based").sum() if "Analysis Source" in os_df_now.columns else 0
                ai_db = (db_df_now.get("Analysis Source", pd.Series(dtype=str)) == "OpenAI").sum() if "Analysis Source" in db_df_now.columns else 0
                rb_db = (db_df_now.get("Analysis Source", pd.Series(dtype=str)) == "Rule-based").sum() if "Analysis Source" in db_df_now.columns else 0
                total_ai = ai_os + ai_db; total_rb = rb_os + rb_db
                if total_ai > 0:
                    st.info(f"🤖 **OpenAI generated {total_ai} Final Recommendations** (OS: {ai_os}, DB: {ai_db}) "
                            f"| 📐 Rule-based fallback: {total_rb} rows")
                else:
                    st.warning("⚠️ OpenAI did not respond — all rows used rule-based analysis. Check API quota.")

            # Verdict summary
            verdicts = ["CRITICAL","UPGRADE NOW","EXTEND + PLAN","REPLACE","CLOUD MIGRATE","MONITOR"]
            fv_col = "Final Verdict"
            if fv_col in os_df_now.columns or fv_col in db_df_now.columns:
                st.subheader("📊 Final Verdict Summary")
                vcols = st.columns(6)
                for i, v in enumerate(verdicts):
                    os_c = int((os_df_now.get(fv_col, pd.Series(dtype=str)).str.upper().str.startswith(v.split()[0])).sum()) if fv_col in os_df_now.columns else 0
                    db_c = int((db_df_now.get(fv_col, pd.Series(dtype=str)).str.upper().str.startswith(v.split()[0])).sum()) if fv_col in db_df_now.columns else 0
                    with vcols[i]: st.metric(v, os_c+db_c, f"OS:{os_c} DB:{db_c}")

            # Guiding principles used
            principles = st.session_state.get("a5_principles", [])
            if principles:
                with st.expander(f"⚖️ {len(principles)} Guiding Principles used", expanded=False):
                    for gp in principles:
                        st.markdown(f"**{gp.get('code','?')}: {gp.get('title','?')}** — {gp.get('rule','')}")

            st.divider()
            col_chat, col_reset = st.columns(2)
            with col_chat:
                if st.button("💬 Continue conversation / refine policy", width="stretch"):
                    st.session_state.a5_status   = "chatting"
                    st.session_state.a5_os_done  = False
                    st.session_state.a5_db_done  = False
                    st.session_state.a5_ws_done  = False
                    st.session_state.a5_as_done  = False
                    st.session_state.a5_fw_done  = False
                    st.rerun()
            with col_reset:
                if st.button("🔄 Start fresh with new policy", width="stretch"):
                    PolicyAnalysisAgent.reset()
                    st.rerun()




# ── Persist Strategist survey data on every render ────────────────────────────
try:
    _save_strategist_state()
except Exception:
    pass

# ── Downloads ──────────────────────────────────────────────────────────────────
st.divider()
dl_col1, dl_col2 = st.columns(2)
with dl_col1:
    ts_str = datetime.now().strftime("%d %b %Y %H:%M UTC")
    excel_bytes = export_to_excel(
        st.session_state.os_df,
        st.session_state.db_df,
        generated_at=ts_str,
        principles=st.session_state.get("a5_principles", []),
        costs=st.session_state.get("a5_costs", {}),
        version_history=VersionGuardianAgent.get_history()
    )
    fname = f"INFY_Version_Tracker_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    os_n, db_n = len(st.session_state.os_df), len(st.session_state.db_df)
    st.download_button(
        label=f"📥 Download Excel — OS: {os_n} · DB: {db_n} entries",
        data=excel_bytes, file_name=fname,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        width="stretch", type="primary"
    )
    st.caption(f"📁 {fname}")

with dl_col2:
    try:
        pdf_bytes = generate_pdf_report(
            st.session_state.os_df,
            st.session_state.db_df,
            principles=st.session_state.get("a5_principles", []),
        )
        pdf_fname = f"INFY_Executive_Report_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
        st.download_button(
            label=f"📄 Download PDF Executive Report",
            data=pdf_bytes, file_name=pdf_fname,
            mime="application/pdf",
            width="stretch"
        )
        st.caption(f"📄 {pdf_fname}")
    except Exception:
        st.caption("PDF export requires fpdf2: `pip install fpdf2`")

st.markdown(
    "<p style='text-align:center;color:#94A3B8;font-size:0.72rem;margin-top:1.5rem;'>"
    "INFY Migration Reference Tracker · Infosys Enterprise Architecture · "
    "Powered by OpenAI GPT · gpt-4o-mini (A1/A2/A5) · gpt-4o-mini (A2/A5) · "
    "All data from AI knowledge base + internet verification</p>",
    unsafe_allow_html=True
)

"""
Excel Export Utility
Generates a polished .xlsx with OS Versions, DB Versions, and Summary sheets.
All data comes from live-fetched dataframes — no hardcoded content.
"""

import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# ── Colour constants ─────────────────────────────────────────────────────────
INFY_NAVY     = "001F5B"
INFY_BLUE     = "003087"
HDR_FONT      = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
HDR_FILL      = PatternFill("solid", fgColor=INFY_BLUE)
TITLE_FILL    = PatternFill("solid", fgColor=INFY_NAVY)
ALT_FILL      = PatternFill("solid", fgColor="EEF2F7")
WHITE_FILL    = PatternFill("solid", fgColor="FFFFFF")
REC_FILL      = PatternFill("solid", fgColor="FFF9C4")

STATUS_FILL = {
    "end of life":    PatternFill("solid", fgColor="FFCDD2"),
    "expiring soon":  PatternFill("solid", fgColor="FFE0B2"),
    "supported":      PatternFill("solid", fgColor="C8E6C9"),
    "future":         PatternFill("solid", fgColor="E3F2FD"),
    "active":         PatternFill("solid", fgColor="C8E6C9"),
}

THIN = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)


def _write_sheet(ws, df: pd.DataFrame, title: str,
                 status_col: str = None, rec_col: str = "Recommendation"):
    cols = list(df.columns)

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols))
    tc = ws.cell(row=1, column=1, value=title)
    tc.font      = Font(bold=True, size=13, color="FFFFFF", name="Calibri")
    tc.fill      = TITLE_FILL
    tc.alignment = CENTER
    ws.row_dimensions[1].height = 28

    # Header row
    for ci, col in enumerate(cols, 1):
        c = ws.cell(row=2, column=ci, value=col)
        c.font = HDR_FONT; c.fill = HDR_FILL
        c.alignment = CENTER; c.border = THIN
    ws.row_dimensions[2].height = 24

    # Data rows
    for ri, (_, row) in enumerate(df.iterrows(), 3):
        base_fill = ALT_FILL if ri % 2 == 0 else WHITE_FILL
        for ci, col in enumerate(cols, 1):
            val = row[col]
            val = "" if pd.isna(val) else str(val)
            c = ws.cell(row=ri, column=ci, value=val)
            c.border = THIN
            c.alignment = LEFT if col in (rec_col, "Notes") else CENTER

            if col == rec_col:
                c.fill = REC_FILL
                c.font = Font(name="Calibri", size=10)
            elif status_col and col == status_col:
                c.fill = STATUS_FILL.get(val.lower(), base_fill)
                c.font = Font(bold=True, name="Calibri", size=10)
            else:
                c.fill = base_fill
                c.font = Font(name="Calibri", size=10)
        ws.row_dimensions[ri].height = 18

    # Column widths
    for ci, col in enumerate(cols, 1):
        cl = get_column_letter(ci)
        if col in (rec_col, "Notes"):
            ws.column_dimensions[cl].width = 55
        elif any(k in col for k in ("Date", "End", "Support")):
            ws.column_dimensions[cl].width = 22
        elif col in ("Upgrade", "Replace"):
            ws.column_dimensions[cl].width = 10
        elif col in ("OS Version", "Database", "Database Version"):
            ws.column_dimensions[cl].width = 32
        else:
            max_len = max(len(col),
                          df[col].astype(str).str.len().max() if len(df) else 0)
            ws.column_dimensions[cl].width = min(max_len + 4, 38)

    ws.freeze_panes = "A3"


def _reorder_os_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Put Recommendation right after OS Version so it's column B."""
    if df.empty or "Recommendation" not in df.columns:
        return df
    cols = list(df.columns)
    cols.remove("Recommendation")
    # Insert after OS Version
    insert_at = cols.index("OS Version") + 1 if "OS Version" in cols else 1
    cols.insert(insert_at, "Recommendation")
    return df[cols]


def _reorder_db_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Put Recommendation right after Database Version so it's column B."""
    if df.empty or "Recommendation" not in df.columns:
        return df
    cols = list(df.columns)
    cols.remove("Recommendation")
    # Insert after Database Version (or first col)
    insert_at = cols.index("Database Version") + 1 if "Database Version" in cols else 1
    cols.insert(insert_at, "Recommendation")
    return df[cols]


def _merge_db_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Merge 'Database' and 'Version' into a single 'Database Version' column."""
    if df.empty:
        return df
    df = df.copy()
    # Create merged column
    df.insert(0, "Database Version",
              df["Database"].fillna("").str.strip() + " " +
              df["Version"].fillna("").str.strip())
    # Drop original separate columns
    df.drop(columns=["Database", "Version"], inplace=True)
    return df


def export_to_excel(os_df: pd.DataFrame, db_df: pd.DataFrame,
                    generated_at: str = None,
                    principles: list = None,
                    costs: dict = None,
                    version_history: list = None) -> bytes:
    """Build and return formatted .xlsx as bytes."""
    wb  = Workbook()
    ts  = generated_at or datetime.now().strftime("%d %b %Y %H:%M UTC")

    # ── Sheet 1 — Final Recommendations (Agent 5) — shown first if available ──
    has_final = (
        "Final Recommendation" in os_df.columns and
        (os_df["Final Recommendation"] != "").any()
    ) or (
        "Final Recommendation" in db_df.columns and
        (db_df["Final Recommendation"] != "").any()
    )

    if has_final:
        ws_final = wb.active
        ws_final.title = "Final Recommendations"

        # Build combined summary: version name + agent2 rec + final rec + verdict
        rows_os, rows_db = [], []
        if "Final Recommendation" in os_df.columns:
            for _, r in os_df.iterrows():
                if r.get("Final Recommendation",""):
                    rows_os.append({
                        "Category":             "OS",
                        "Name":                 r.get("OS Version",""),
                        "Support Ends":         r.get("Extended/LTSC Support End","") or r.get("Mainstream/Full Support End",""),
                        "Agent 2 Recommendation": r.get("Recommendation",""),
                        "Final Recommendation": r.get("Final Recommendation",""),
                        "Final Verdict":        r.get("Final Verdict",""),
                    })
        if "Final Recommendation" in db_df.columns:
            for _, r in db_df.iterrows():
                if r.get("Final Recommendation",""):
                    rows_db.append({
                        "Category":             "DB",
                        "Name":                 f"{r.get('Database','')} {r.get('Version','')}",
                        "Support Ends":         r.get("Extended Support End","") or r.get("Mainstream / Premier End",""),
                        "Agent 2 Recommendation": r.get("Recommendation",""),
                        "Final Recommendation": r.get("Final Recommendation",""),
                        "Final Verdict":        r.get("Final Verdict",""),
                    })

        combined = pd.DataFrame(rows_os + rows_db)
        if not combined.empty:
            # Sort by verdict severity
            verdict_order = {"CRITICAL":0,"UPGRADE NOW":1,"EXTEND + PLAN":2,
                             "REPLACE":3,"CLOUD MIGRATE":4,"MONITOR":5}
            combined["_sort"] = combined["Final Verdict"].map(
                lambda v: verdict_order.get(v, 9))
            combined = combined.sort_values("_sort").drop(columns=["_sort"])

            _write_sheet(ws_final, combined,
                         title=f"Final Recommendations — INFY Migration Reference  |  {ts}",
                         status_col="Final Verdict", rec_col="Final Recommendation")

            # Wider columns for this sheet
            ws_final.column_dimensions["A"].width = 10   # Category
            ws_final.column_dimensions["B"].width = 32   # Name
            ws_final.column_dimensions["C"].width = 18   # Support Ends
            ws_final.column_dimensions["D"].width = 52   # Agent 2 Rec
            ws_final.column_dimensions["E"].width = 52   # Final Rec
            ws_final.column_dimensions["F"].width = 16   # Verdict

        # OS and DB sheets second and third
        ws_os = wb.create_sheet("OS Versions")
        ws_db = wb.create_sheet("DB Versions")
    else:
        # No final recs yet — OS is first sheet
        ws_os = wb.active
        ws_os.title = "OS Versions"
        ws_db = wb.create_sheet("DB Versions")

    # ── OS Versions ────────────────────────────────────────────────────────────
    _write_sheet(ws_os, _reorder_os_columns(os_df),
                 title=f"OS Versions — INFY Migration Reference  |  {ts}",
                 status_col=None, rec_col="Recommendation")

    # ── DB Versions ────────────────────────────────────────────────────────────
    db_merged = _reorder_db_columns(_merge_db_columns(db_df))
    _write_sheet(ws_db, db_merged,
                 title=f"DB Versions — INFY Migration Reference  |  {ts}",
                 status_col="Status", rec_col="Recommendation")

    # ── Summary ────────────────────────────────────────────────────────────────
    ws_s = wb.create_sheet("Summary")
    rows = [
        ("", ""),
        ("INFY Migration Version Lifecycle Tracker", ""),
        ("", ""),
        ("Generated on",        ts),
        ("OS Versions (rows)",  str(len(os_df))),
        ("DB Versions (rows)",  str(len(db_df))),
        ("Project Window",      "1 April 2026 → 30 June 2028"),
        ("", ""),
        ("Agent 1", "Checks internet for lifecycle date changes — gpt-4o-mini-search-preview"),
        ("Agent 2", "Generates expert technical recommendations per row — gpt-4o-mini"),
        ("Agent 3", "14-day refresh monitor — seeks permission before re-running"),
        ("Agent 4", "Version Guardian — snapshots data before every refresh"),
        ("Agent 5", "Conversational policy interview → Guiding Principles → Final Recommendations sheet"),
        ("", ""),
        ("AI Model",    "OpenAI gpt-4o-mini + gpt-4o-mini-search-preview"),
        ("Data Source", "AI knowledge base + live web verification"),
    ]

    ws_s.column_dimensions["A"].width = 30
    ws_s.column_dimensions["B"].width = 80

    for ri, (label, value) in enumerate(rows, 1):
        ca = ws_s.cell(row=ri, column=1, value=label)
        cb = ws_s.cell(row=ri, column=2, value=value)
        if ri == 2:
            ca.font = Font(bold=True, size=14, color=INFY_NAVY, name="Calibri")
            ws_s.merge_cells(f"A{ri}:B{ri}")
            ca.alignment = CENTER
        elif label in ("Agent 1","Agent 2","Agent 3","Agent 4","Agent 5"):
            ca.font = Font(bold=True, size=11, color=INFY_BLUE, name="Calibri")
            cb.font = Font(size=11, name="Calibri")
        elif label:
            ca.font = Font(bold=True, size=11, name="Calibri")
            cb.font = Font(size=11, name="Calibri")
        ws_s.row_dimensions[ri].height = 20

    # ── Guiding Principles ─────────────────────────────────────────────────────
    if principles:
        ws_gp = wb.create_sheet("Guiding Principles")
        ws_gp.column_dimensions["A"].width = 10
        ws_gp.column_dimensions["B"].width = 28
        ws_gp.column_dimensions["C"].width = 55
        ws_gp.column_dimensions["D"].width = 48
        ws_gp.column_dimensions["E"].width = 14
        for ci, h in enumerate(["Code","Title","Rule","Trigger","Category"], 1):
            c = ws_gp.cell(row=1, column=ci, value=h)
            c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = CENTER; c.border = THIN
        for ri, gp in enumerate(principles, 2):
            fill = ALT_FILL if ri % 2 == 0 else WHITE_FILL
            for ci, key in enumerate(["code","title","rule","trigger","category"], 1):
                c = ws_gp.cell(row=ri, column=ci, value=gp.get(key,""))
                c.font = Font(name="Calibri", size=10)
                c.fill = fill; c.border = THIN
                c.alignment = LEFT if ci > 1 else CENTER
            ws_gp.row_dimensions[ri].height = 16

    # ── Vendor Cost Data ───────────────────────────────────────────────────────
    if costs:
        ws_cost = wb.create_sheet("Vendor Cost Data")
        ws_cost.column_dimensions["A"].width = 28
        ws_cost.column_dimensions["B"].width = 88
        for ci, h in enumerate(["Vendor","Cost Summary"], 1):
            c = ws_cost.cell(row=1, column=ci, value=h)
            c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = CENTER; c.border = THIN
        for ri, (vendor, summary) in enumerate(costs.items(), 2):
            fill = ALT_FILL if ri % 2 == 0 else WHITE_FILL
            cv = ws_cost.cell(row=ri, column=1, value=vendor)
            cs = ws_cost.cell(row=ri, column=2, value=summary)
            for c in (cv, cs):
                c.font = Font(name="Calibri", size=10)
                c.fill = fill; c.border = THIN; c.alignment = LEFT
            ws_cost.row_dimensions[ri].height = 18

    # ── Version History Snapshots ──────────────────────────────────────────────
    if version_history:
        def _safe_sheet(name: str, suffix: str) -> str:
            safe = name.replace(":", "").replace("\\","").replace("/","") \
                       .replace("?","").replace("*","").replace("[","").replace("]","")
            max_base = 31 - len(suffix) - 1
            return f"{safe[:max_base]} {suffix}"

        for snap in version_history:
            label = snap["label"]
            ws_snap_os = wb.create_sheet(_safe_sheet(label, "OS"))
            _write_sheet(ws_snap_os, _reorder_os_columns(snap["os_df"]),
                         title=f"OS Snapshot — {label}",
                         status_col=None, rec_col="Recommendation")
            ws_snap_db = wb.create_sheet(_safe_sheet(label, "DB"))
            _write_sheet(ws_snap_db, _reorder_db_columns(_merge_db_columns(snap["db_df"])),
                         title=f"DB Snapshot — {label}",
                         status_col="Status", rec_col="Recommendation")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
